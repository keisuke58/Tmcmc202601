#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
validate_monospecies.py — Test validation for monospecies Hamilton ODE + TMCMC.

Checks:
  1. Data loading (St sheet)
  2. ODE trajectories: static c=100 vs dynamic c(t)=10*(1-phi*psi)
  3. Profiled RMSE landscape over sigma grid → find MAP
  4. Prior appropriateness for a₁₁
"""

import os
import sys
import numpy as np

sys.path.insert(0, os.path.dirname(__file__))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'main'))

import jax
jax.config.update('jax_enable_x64', True)
import jax.numpy as jnp
from jax import jit
from functools import partial

from hamilton_monospecies_jax import (
    simulate_monospecies, newton_step_4d, clip4, residual_4d
)

# ── 0. Paper reference values ─────────────────────────────────────────────────
PAPER_SIGMA = {4: 25, 8: 50, 15: 50, 20: 70, 25: 70, 35: 110, 37: 110, 40: 50}
OUR_SIGMA   = {4: 4.25, 8: 10, 15: 25, 20: 50, 25: 100, 35: 110, 37: 115, 40: 40}
N_STEPS_MAP = {4: 4000, 8: 2000, 15: 1000, 20: 500, 25: 300, 35: 200, 37: 200, 40: 450}
TIME_TO_STEP = 10.0  # 1 day = 10 steps

# ── 1. Data loading ───────────────────────────────────────────────────────────

def load_data(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['St']
    temps = [4, 8, 15, 20, 25, 35, 37, 40]
    data = {}
    for i, temp in enumerate(temps):
        col_start = i * 3
        times_all, cfu_all = [], []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
            t_cell = row[col_start]
            c_cell = row[col_start + 1]
            if t_cell.value is not None and c_cell.value is not None:
                try:
                    times_all.append(float(t_cell.value))
                    cfu_all.append(float(c_cell.value))
                except (TypeError, ValueError):
                    pass
        unique_times = sorted(set(times_all))
        avg_cfu = [np.mean([cfu_all[j] for j in range(len(times_all))
                            if times_all[j] == t]) for t in unique_times]
        data[temp] = {'time': np.array(unique_times), 'cfu_mean': np.array(avg_cfu)}
    return data


# ── 2. Dynamic-c ODE ─────────────────────────────────────────────────────────

def simulate_dynamic_c(sigma, n_steps, dt=1e-4, phi_init=0.1,
                       Kp=1e-4, eta=1.0, eta_phi=1.0, c_scale=10.0, n_newton=6):
    """
    c(t) = c_scale * (1 - phi*psi)  — dynamic nutrient coupling.
    Uses Python loop (JAX can't JIT dynamic c via fori_loop easily).
    """
    sigma = float(sigma)
    dt = float(dt)

    phi0_init = 1.0 - phi_init
    psi_init = 0.999
    g = jnp.array([phi_init, phi0_init, psi_init, 0.0])

    phi_list = [phi_init]
    psi_list = [psi_init]

    for _ in range(n_steps):
        phi, phi0, psi, gamma = float(g[0]), float(g[1]), float(g[2]), float(g[3])
        c_t = c_scale * (1.0 - phi * psi)
        g = newton_step_4d(g, jnp.float64(sigma), jnp.float64(dt),
                           jnp.float64(Kp), jnp.float64(eta), jnp.float64(eta_phi),
                           jnp.float64(c_t), jnp.float64(0.0), jnp.float64(0.0), n_newton)
        phi_list.append(float(g[0]))
        psi_list.append(float(g[2]))

    phi_hist = np.array(phi_list)
    psi_hist = np.array(psi_list)
    phibar_hist = phi_hist * psi_hist
    return phi_hist, psi_hist, phibar_hist


# ── 3. Profiled RMSE ─────────────────────────────────────────────────────────

def profiled_rmse(phibar, exp_time, exp_cfu, n_steps):
    """Affine map phibar → cfu, return RMSE."""
    step_idx = np.clip((exp_time * TIME_TO_STEP).astype(int), 0, n_steps)
    sim_vals = phibar[step_idx]
    X = np.column_stack([sim_vals, np.ones_like(sim_vals)])
    coeff, _, _, _ = np.linalg.lstsq(X, exp_cfu, rcond=None)
    predicted = X @ coeff
    return float(np.sqrt(np.mean((predicted - exp_cfu)**2))), coeff


def profiled_loglik(phibar, exp_time, exp_cfu, sigma_obs, n_steps):
    """Profiled Gaussian log-likelihood (matches estimate_monospecies_tmcmc.py)."""
    step_idx = np.clip((exp_time * TIME_TO_STEP).astype(int), 0, n_steps)
    sim_vals = phibar[step_idx]
    X = np.column_stack([sim_vals, np.ones_like(sim_vals)])
    coeff, _, _, _ = np.linalg.lstsq(X, exp_cfu, rcond=None)
    predicted = X @ coeff
    residuals = (predicted - exp_cfu) / sigma_obs
    return float(-0.5 * np.sum(residuals**2) - np.sum(np.log(sigma_obs)))


# ── 4. Main validation ────────────────────────────────────────────────────────

def validate_all(data):
    temps = [4, 8, 15, 20, 25, 35, 37, 40]
    sigma_grid = np.concatenate([
        np.linspace(1, 10, 10),
        np.linspace(10, 50, 20),
        np.linspace(50, 200, 30),
        np.linspace(200, 500, 10),
    ])
    sigma_grid = np.unique(sigma_grid)

    print("\n" + "="*100)
    print("TEST 1: Data loading")
    print("="*100)
    for temp in temps:
        d = data[temp]
        print(f"  {temp:>2d}°C: n_obs={len(d['time'])}, "
              f"time=[{d['time'][0]:.1f}..{d['time'][-1]:.1f}], "
              f"cfu=[{d['cfu_mean'].min():.3f}..{d['cfu_mean'].max():.3f}]")

    print("\n" + "="*100)
    print("TEST 2: ODE trajectory (static c=100 vs dynamic c=10*(1-phi*psi))")
    print("  Using OUR sigma values from felix_monospecies_dynamic_c_params.md")
    print("="*100)
    print(f"{'Temp':>4s} | {'σ_ours':>7s} | {'static phibar_final':>18s} | {'dynamic phibar_final':>20s} | {'Δ%':>6s}")
    print("-"*70)
    for temp in temps:
        sigma = OUR_SIGMA[temp]
        n_steps = N_STEPS_MAP[temp]
        _, _, pb_static = simulate_monospecies(sigma, n_steps, c=100.0, alpha=0.0, b=0.0)
        _, _, pb_dyn    = simulate_dynamic_c(sigma, n_steps, c_scale=10.0)
        pb_s = float(pb_static[-1])
        pb_d = float(pb_dyn[-1])
        delta_pct = (pb_d - pb_s) / abs(pb_s) * 100 if abs(pb_s) > 1e-10 else 0.0
        print(f"  {temp:>2d}°C | {sigma:>7.2f} | {pb_s:>18.6f} | {pb_d:>20.6f} | {delta_pct:>+6.1f}%")

    print("\n" + "="*100)
    print("TEST 3: Profiled RMSE at paper σ vs our σ (static c=100)")
    print("="*100)
    print(f"{'Temp':>4s} | {'σ_paper':>8s} | {'RMSE_paper':>10s} | {'σ_ours':>7s} | {'RMSE_ours':>10s}")
    print("-"*60)
    for temp in temps:
        d = data[temp]
        n_steps = N_STEPS_MAP[temp]
        # Paper σ
        _, _, pb_p = simulate_monospecies(PAPER_SIGMA[temp], n_steps, c=100.0)
        rmse_p, _ = profiled_rmse(np.array(pb_p), d['time'], d['cfu_mean'], n_steps)
        # Our σ
        _, _, pb_o = simulate_monospecies(OUR_SIGMA[temp], n_steps, c=100.0)
        rmse_o, _ = profiled_rmse(np.array(pb_o), d['time'], d['cfu_mean'], n_steps)
        print(f"  {temp:>2d}°C | {PAPER_SIGMA[temp]:>8d} | {rmse_p:>10.6f} | "
              f"{OUR_SIGMA[temp]:>7.2f} | {rmse_o:>10.6f}")

    print("\n" + "="*100)
    print("TEST 4: Profiled RMSE at paper σ vs our σ (DYNAMIC c)")
    print("="*100)
    print(f"{'Temp':>4s} | {'σ_paper':>8s} | {'RMSE_paper_dyn':>14s} | {'σ_ours':>7s} | {'RMSE_ours_dyn':>13s}")
    print("-"*70)
    for temp in temps:
        d = data[temp]
        n_steps = N_STEPS_MAP[temp]
        _, _, pb_p = simulate_dynamic_c(PAPER_SIGMA[temp], n_steps)
        rmse_p, _ = profiled_rmse(np.array(pb_p), d['time'], d['cfu_mean'], n_steps)
        _, _, pb_o = simulate_dynamic_c(OUR_SIGMA[temp], n_steps)
        rmse_o, _ = profiled_rmse(np.array(pb_o), d['time'], d['cfu_mean'], n_steps)
        print(f"  {temp:>2d}°C | {PAPER_SIGMA[temp]:>8d} | {rmse_p:>14.6f} | "
              f"{OUR_SIGMA[temp]:>7.2f} | {rmse_o:>13.6f}")

    print("\n" + "="*100)
    print("TEST 5: Likelihood landscape scan (static c=100) — finds MAP σ per condition")
    print("="*100)
    print(f"{'Temp':>4s} | {'MAP_σ':>8s} | {'MAP_logL':>10s} | {'σ_ours':>7s} | {'logL_ours':>10s} | "
          f"{'σ_paper':>8s} | {'logL_paper':>11s} | {'σ_range_95%':>15s}")
    print("-"*105)
    sigma_obs_default = 0.2

    grid_results = {}
    for temp in temps:
        d = data[temp]
        n_steps = N_STEPS_MAP[temp]
        sigma_obs = np.full(len(d['time']), sigma_obs_default)

        logL_grid = []
        for sig in sigma_grid:
            try:
                _, _, pb = simulate_monospecies(float(sig), n_steps, c=100.0)
                ll = profiled_loglik(np.array(pb), d['time'], d['cfu_mean'], sigma_obs, n_steps)
            except Exception:
                ll = -1e10
            logL_grid.append(ll)
        logL_grid = np.array(logL_grid)

        # MAP from grid
        best_idx = np.argmax(logL_grid)
        map_sigma = sigma_grid[best_idx]
        map_logL = logL_grid[best_idx]

        # logL at our σ
        _, _, pb_o = simulate_monospecies(OUR_SIGMA[temp], n_steps, c=100.0)
        ll_ours = profiled_loglik(np.array(pb_o), d['time'], d['cfu_mean'], sigma_obs, n_steps)

        # logL at paper σ
        _, _, pb_p = simulate_monospecies(PAPER_SIGMA[temp], n_steps, c=100.0)
        ll_paper = profiled_loglik(np.array(pb_p), d['time'], d['cfu_mean'], sigma_obs, n_steps)

        # 95% CI from grid (posterior ∝ exp(logL), uniform prior)
        logL_rel = logL_grid - map_logL
        posterior = np.exp(logL_rel)
        posterior /= (posterior.sum() + 1e-30)
        cdf = np.cumsum(posterior)
        lo_idx = np.searchsorted(cdf, 0.025)
        hi_idx = np.searchsorted(cdf, 0.975)
        ci_lo = sigma_grid[max(0, lo_idx)]
        ci_hi = sigma_grid[min(len(sigma_grid)-1, hi_idx)]

        grid_results[temp] = {'logL': logL_grid, 'map_sigma': map_sigma, 'map_logL': map_logL}
        print(f"  {temp:>2d}°C | {map_sigma:>8.2f} | {map_logL:>10.2f} | "
              f"{OUR_SIGMA[temp]:>7.2f} | {ll_ours:>10.2f} | "
              f"{PAPER_SIGMA[temp]:>8d} | {ll_paper:>11.2f} | "
              f"[{ci_lo:>5.1f}, {ci_hi:>5.1f}]")

    print("\n" + "="*100)
    print("TEST 6: Likelihood landscape scan (DYNAMIC c) — finds MAP σ per condition")
    print("="*100)
    print(f"{'Temp':>4s} | {'MAP_σ':>8s} | {'MAP_logL':>10s} | {'σ_ours':>7s} | {'logL_ours':>10s} | "
          f"{'σ_paper':>8s} | {'logL_paper':>11s}")
    print("-"*80)
    sigma_grid_coarse = np.concatenate([
        np.linspace(1, 20, 8), np.linspace(20, 100, 10), np.linspace(100, 300, 8)
    ])
    sigma_grid_coarse = np.unique(sigma_grid_coarse)

    for temp in temps:
        d = data[temp]
        n_steps = N_STEPS_MAP[temp]
        sigma_obs = np.full(len(d['time']), sigma_obs_default)

        logL_grid = []
        for sig in sigma_grid_coarse:
            try:
                _, _, pb = simulate_dynamic_c(float(sig), n_steps)
                ll = profiled_loglik(np.array(pb), d['time'], d['cfu_mean'], sigma_obs, n_steps)
            except Exception:
                ll = -1e10
            logL_grid.append(ll)
        logL_grid = np.array(logL_grid)

        best_idx = np.argmax(logL_grid)
        map_sigma = sigma_grid_coarse[best_idx]
        map_logL = logL_grid[best_idx]

        _, _, pb_o = simulate_dynamic_c(OUR_SIGMA[temp], n_steps)
        ll_ours = profiled_loglik(np.array(pb_o), d['time'], d['cfu_mean'], sigma_obs, n_steps)

        _, _, pb_p = simulate_dynamic_c(PAPER_SIGMA[temp], n_steps)
        ll_paper = profiled_loglik(np.array(pb_p), d['time'], d['cfu_mean'], sigma_obs, n_steps)

        print(f"  {temp:>2d}°C | {map_sigma:>8.2f} | {map_logL:>10.2f} | "
              f"{OUR_SIGMA[temp]:>7.2f} | {ll_ours:>10.2f} | "
              f"{PAPER_SIGMA[temp]:>8d} | {ll_paper:>11.2f}")

    print("\n" + "="*100)
    print("TEST 7: Prior sensitivity for a₁₁")
    print("  Current prior: Uniform[1, 500]")
    print("  Felix params: min=4.25 (4°C), max=115 (37°C)")
    print("  Assessment:")
    our_vals = list(OUR_SIGMA.values())
    print(f"    Our σ range: [{min(our_vals)}, {max(our_vals)}]")
    print(f"    Prior [1, 500] covers all values ✓")
    print(f"    Log-uniform [log 1, log 500] = [0, 6.2] might be better")
    print(f"    (values span 4.25→115, ratio ~27×, suggests log-scale prior)")
    print("="*100)


if __name__ == '__main__':
    import time as _t
    xlsx_path = os.path.join(os.path.dirname(__file__), 'raw data.xlsx')
    print(f"Loading data from: {xlsx_path}")
    data = load_data(xlsx_path)

    t0 = _t.time()
    validate_all(data)
    print(f"\nTotal validation time: {_t.time()-t0:.1f}s")
