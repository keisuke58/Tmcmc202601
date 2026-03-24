# -*- coding: utf-8 -*-
"""
hamilton_monospecies.py — Felix Klempt's Hamilton ODE (Numba), monospecies version.

Exact reproduction of Felix's Mathematica code (Biofilm-TSM-Mathematica-2025-01-17):

Lagrangian:
  ℋ = c(t)·Φ·A·Φ - α(t)·Ψ·B·Ψ + Σ Penalty_i + γ(Σφ_i + φ₀ - 1)

where:
  Φ = (φ₁ψ₁, ...), A = -(1/2)·[[a₁₁,...]], B = -(1/2)·diag(b₁₁,...)
  Penalty_i = η_i · Kp · 1/(φ_i²(1-φ_i)²)      [value, NOT derivative]
  Penalty_ψi = η_i · Kp · 1/(ψ_i²(1-ψ_i)²)
  Penalty_φ₀ = Kp · 1/(φ₀²(1-φ₀)²)

Dissipation:
  Δ_rd = (1/2)·Φ̇·η·Φ̇ + (1/2)·φ̇·η_φ·φ̇
  Δ_0  = (1/2)·φ₀̇²

Residuals (Euler-Lagrange):
  Q_φ[i] = (∂ℋ/∂φ[i] + ∂Δ_rd/∂φ̇[i]) / η_i
  Q_ψ[i] = ∂ℋ/∂ψ[i] + ∂Δ_rd/∂ψ̇[i]
  Q_φ₀  = ∂ℋ/∂φ₀ + ∂Δ₀/∂φ̇₀
  Q_γ   = ∂ℋ/∂γ = Σφ_i + φ₀ - 1

Monospecies: N=1, state = [φ, φ₀, ψ, γ].
"""

import numpy as np
from numba import njit
from scipy.optimize import differential_evolution
import time as _time


@njit(cache=True)
def _clip(val, lo, hi):
    if val < lo: return lo
    if val > hi: return hi
    return val


@njit(cache=True)
def _residual_felix(phi, phi0, psi, gamma,
                    phi_old, phi0_old, psi_old,
                    a11, b11, dt, Kp, eta, eta_phi, c, alpha):
    """
    Felix's Hamilton residual for monospecies (N=1).

    ℋ = c · Φ·A·Φ - α · ψ·B·ψ + Penalty_φ + Penalty_ψ + Penalty_φ₀ + γ(φ + φ₀ - 1)

    where Φ = φ·ψ (scalar), A = -(1/2)·a₁₁, B = -(1/2)·b₁₁

    So: c·Φ·A·Φ = c · (φψ) · (-(1/2)a₁₁) · (φψ) = -(c/2)·a₁₁·(φψ)²
        α·ψ·B·ψ = α · ψ · (-(1/2)b₁₁) · ψ = -(α/2)·b₁₁·ψ²

    Penalty_φ  = η · Kp / (φ²(1-φ)²)
    Penalty_ψ  = η · Kp / (ψ²(1-ψ)²)
    Penalty_φ₀ = Kp / (φ₀²(1-φ₀)²)

    ∂ℋ/∂φ = c · (-(1/2)a₁₁) · 2·(φψ)·ψ + ∂Penalty_φ/∂φ + γ
           = -c·a₁₁·φ·ψ² + η·Kp·∂/∂φ[1/(φ²(1-φ)²)] + γ

    ∂Δ_rd/∂φ̇ = η · (φ̇ψ + φψ̇)·ψ  [from Φ̇·η·Φ̇ chain]
              Actually: Δ_rd = (1/2)·η·Φ̇² + (1/2)·η_φ·φ̇²
              Φ̇ = φ̇ψ + φψ̇
              ∂Δ_rd/∂φ̇ = η·Φ̇·ψ + η_φ·φ̇  (since ∂Φ̇/∂φ̇ = ψ)

    Q_φ = (∂ℋ/∂φ + ∂Δ_rd/∂φ̇) / η
    """
    phidot = (phi - phi_old) / dt
    phi0dot = (phi0 - phi0_old) / dt
    psidot = (psi - psi_old) / dt

    Phi = phi * psi
    Phidot = phidot * psi + phi * psidot

    # ∂/∂φ of Penalty_φ = η·Kp/(φ²(1-φ)²)
    # = η·Kp · (-2φ + 2φ² - 2φ + 2φ²) / (φ⁴(1-φ)⁴)
    # = η·Kp · (2-4φ) / ((φ-1)³ φ³)  ... standard form
    dPen_phi = eta * Kp * (2.0 - 4.0 * phi) / ((phi - 1.0)**3 * phi**3)

    # ∂/∂ψ of Penalty_ψ = η·Kp/(ψ²(1-ψ)²)
    dPen_psi = eta * Kp * (2.0 - 4.0 * psi) / ((psi - 1.0)**3 * psi**3)

    # ∂/∂φ₀ of Penalty_φ₀ = Kp/(φ₀²(1-φ₀)²)
    dPen_phi0 = Kp * (2.0 - 4.0 * phi0) / ((phi0 - 1.0)**3 * phi0**3)

    # ∂ℋ/∂φ = -c·a₁₁·φ·ψ² + dPen_phi + γ
    dH_dphi = -c * a11 * phi * psi**2 + dPen_phi + gamma

    # ∂ℋ/∂ψ = -c·a₁₁·φ²·ψ + α·b₁₁·ψ + dPen_psi
    # (from -(c/2)·a₁₁·(φψ)² → ∂/∂ψ = -c·a₁₁·φ²·ψ)
    # (from -(-α/2)·b₁₁·ψ² = (α/2)·b₁₁·ψ² → ∂/∂ψ = α·b₁₁·ψ)
    dH_dpsi = -c * a11 * phi**2 * psi + alpha * b11 * psi + dPen_psi

    # ∂ℋ/∂φ₀ = dPen_phi0 + γ
    dH_dphi0 = dPen_phi0 + gamma

    # ∂ℋ/∂γ = φ + φ₀ - 1
    dH_dgamma = phi + phi0 - 1.0

    # Dissipation derivatives
    # Δ_rd = (1/2)·η·Φ̇² + (1/2)·η_φ·φ̇²
    # ∂Δ_rd/∂φ̇ = η·Φ̇·ψ + η_φ·φ̇
    dDelta_dphidot = eta * Phidot * psi + eta_phi * phidot

    # ∂Δ_rd/∂ψ̇ = η·Φ̇·φ
    dDelta_dpsidot = eta * Phidot * phi

    # Δ₀ = (1/2)·φ₀̇²
    # ∂Δ₀/∂φ₀̇ = φ₀̇
    dDelta0_dphi0dot = phi0dot

    # Residuals
    # Q_φ = (∂ℋ/∂φ + ∂Δ_rd/∂φ̇) / η_i  ← divided by η!
    Q_phi = (dH_dphi + dDelta_dphidot) / eta

    # Q_ψ = (∂ℋ/∂ψ + ∂Δ_rd/∂ψ̇) / η_i  ← ALSO divided by η!
    Q_psi = (dH_dpsi + dDelta_dpsidot) / eta

    # Q_φ₀ = ∂ℋ/∂φ₀ + ∂Δ₀/∂φ₀̇
    Q_phi0 = dH_dphi0 + dDelta0_dphi0dot

    # Q_γ = ∂ℋ/∂γ
    Q_gamma = dH_dgamma

    return Q_phi, Q_phi0, Q_psi, Q_gamma


@njit(cache=True)
def _residual_felix_dynamic_c_exogenous(phi, phi0, psi, gamma,
                                       phi_old, phi0_old, psi_old,
                                       a11, b11, dt, Kp, eta, eta_phi,
                                       alpha):
    """
    Felix formulation where c(t) is computed from current state, but treated
    as an exogenous time function (i.e., dc/dphi and dc/dpsi terms are NOT
    included in the variational derivatives).

    c(t) = 10 * (1 - phi*psi)
    """
    phidot = (phi - phi_old) / dt
    phi0dot = (phi0 - phi0_old) / dt
    psidot = (psi - psi_old) / dt

    Phi = phi * psi
    Phidot = phidot * psi + phi * psidot

    # c(t) computed from the current state
    c_dyn = 10.0 * (1.0 - Phi)

    dPen_phi = eta * Kp * (2.0 - 4.0 * phi) / ((phi - 1.0)**3 * phi**3)
    dPen_psi = eta * Kp * (2.0 - 4.0 * psi) / ((psi - 1.0)**3 * psi**3)
    dPen_phi0 = Kp * (2.0 - 4.0 * phi0) / ((phi0 - 1.0)**3 * phi0**3)

    dH_dphi = -c_dyn * a11 * phi * psi**2 + dPen_phi + gamma
    dH_dpsi = -c_dyn * a11 * phi**2 * psi + alpha * b11 * psi + dPen_psi
    dH_dphi0 = dPen_phi0 + gamma

    dH_dgamma = phi + phi0 - 1.0

    dDelta_dphidot = eta * Phidot * psi + eta_phi * phidot
    dDelta_dpsidot = eta * Phidot * phi
    dDelta0_dphi0dot = phi0dot

    Q_phi = (dH_dphi + dDelta_dphidot) / eta
    Q_psi = (dH_dpsi + dDelta_dpsidot) / eta
    Q_phi0 = dH_dphi0 + dDelta0_dphi0dot
    Q_gamma = dH_dgamma

    return Q_phi, Q_phi0, Q_psi, Q_gamma


@njit(cache=True)
def _residual_felix_dynamic_c_exogenous_scale(phi, phi0, psi, gamma,
                                             phi_old, phi0_old, psi_old,
                                             a11, b11, dt, Kp, eta, eta_phi,
                                             c_scale, alpha):
    phidot = (phi - phi_old) / dt
    phi0dot = (phi0 - phi0_old) / dt
    psidot = (psi - psi_old) / dt

    Phi = phi * psi
    Phidot = phidot * psi + phi * psidot

    c_dyn = c_scale * (1.0 - Phi)

    dPen_phi = eta * Kp * (2.0 - 4.0 * phi) / ((phi - 1.0)**3 * phi**3)
    dPen_psi = eta * Kp * (2.0 - 4.0 * psi) / ((psi - 1.0)**3 * psi**3)
    dPen_phi0 = Kp * (2.0 - 4.0 * phi0) / ((phi0 - 1.0)**3 * phi0**3)

    dH_dphi = -c_dyn * a11 * phi * psi**2 + dPen_phi + gamma
    dH_dpsi = -c_dyn * a11 * phi**2 * psi + alpha * b11 * psi + dPen_psi
    dH_dphi0 = dPen_phi0 + gamma

    dH_dgamma = phi + phi0 - 1.0

    dDelta_dphidot = eta * Phidot * psi + eta_phi * phidot
    dDelta_dpsidot = eta * Phidot * phi
    dDelta0_dphi0dot = phi0dot

    Q_phi = (dH_dphi + dDelta_dphidot) / eta
    Q_psi = (dH_dpsi + dDelta_dpsidot) / eta
    Q_phi0 = dH_dphi0 + dDelta0_dphi0dot
    Q_gamma = dH_dgamma

    return Q_phi, Q_phi0, Q_psi, Q_gamma


@njit(cache=True)
def _residual_felix_dynamic_c_state_substituted(phi, phi0, psi, gamma,
                                                  phi_old, phi0_old, psi_old,
                                                  a11, b11, dt, Kp, eta,
                                                  eta_phi, alpha):
    """
    Same c(t) = 10*(1-phi*psi), but with full state-substitution inside the
    Hamilton derivatives (i.e., extra terms from dc/dphi and dc/dpsi are
    included).
    """
    phidot = (phi - phi_old) / dt
    phi0dot = (phi0 - phi0_old) / dt
    psidot = (psi - psi_old) / dt

    Phidot = phidot * psi + phi * psidot

    # Expanded derivative of (-a11/2) * c(phi,psi) * (phi*psi)^2 with
    # c(phi,psi) = 10*(1-phi*psi).
    # d/dphi term: -10*a11*phi*psi^2 + 15*a11*phi^2*psi^3
    dPen_phi = eta * Kp * (2.0 - 4.0 * phi) / ((phi - 1.0)**3 * phi**3)
    dPen_psi = eta * Kp * (2.0 - 4.0 * psi) / ((psi - 1.0)**3 * psi**3)
    dPen_phi0 = Kp * (2.0 - 4.0 * phi0) / ((phi0 - 1.0)**3 * phi0**3)

    dH_dphi = (-10.0 * a11 * phi * psi**2
               + 15.0 * a11 * phi**2 * psi**3) + dPen_phi + gamma

    # d/dpsi term: -10*a11*phi^2*psi + 15*a11*phi^3*psi^2
    dH_dpsi = (-10.0 * a11 * phi**2 * psi
               + 15.0 * a11 * phi**3 * psi**2) + alpha * b11 * psi + dPen_psi

    dH_dphi0 = dPen_phi0 + gamma
    dH_dgamma = phi + phi0 - 1.0

    dDelta_dphidot = eta * Phidot * psi + eta_phi * phidot
    dDelta_dpsidot = eta * Phidot * phi
    dDelta0_dphi0dot = phi0dot

    Q_phi = (dH_dphi + dDelta_dphidot) / eta
    Q_psi = (dH_dpsi + dDelta_dpsidot) / eta
    Q_phi0 = dH_dphi0 + dDelta0_dphi0dot
    Q_gamma = dH_dgamma

    return Q_phi, Q_phi0, Q_psi, Q_gamma


@njit(cache=True)
def _newton_step_felix(phi_old, phi0_old, psi_old, gamma_old,
                       a11, b11, dt, Kp, eta, eta_phi, c, alpha, n_iter):
    eps = 1e-10
    eps_fd = 1e-8

    phi = _clip(phi_old, eps, 1.0 - eps)
    phi0 = _clip(phi0_old, eps, 1.0 - eps)
    psi = _clip(psi_old, eps, 1.0 - eps)
    gamma = _clip(gamma_old, -1e6, 1e6)

    for _ in range(n_iter):
        Q0, Q1, Q2, Q3 = _residual_felix(phi, phi0, psi, gamma,
                                           phi_old, phi0_old, psi_old,
                                           a11, b11, dt, Kp, eta, eta_phi, c, alpha)
        Q_base = np.array([Q0, Q1, Q2, Q3])

        # Numerical Jacobian
        J = np.zeros((4, 4))
        g = np.array([phi, phi0, psi, gamma])
        for j in range(4):
            gp = g.copy()
            gp[j] += eps_fd
            gp[0] = _clip(gp[0], eps, 1.0 - eps)
            gp[1] = _clip(gp[1], eps, 1.0 - eps)
            gp[2] = _clip(gp[2], eps, 1.0 - eps)
            gp[3] = _clip(gp[3], -1e6, 1e6)
            Qp0, Qp1, Qp2, Qp3 = _residual_felix(gp[0], gp[1], gp[2], gp[3],
                                                    phi_old, phi0_old, psi_old,
                                                    a11, b11, dt, Kp, eta, eta_phi, c, alpha)
            J[0, j] = (Qp0 - Q_base[0]) / eps_fd
            J[1, j] = (Qp1 - Q_base[1]) / eps_fd
            J[2, j] = (Qp2 - Q_base[2]) / eps_fd
            J[3, j] = (Qp3 - Q_base[3]) / eps_fd

        # Solve 4x4 with partial pivoting
        rhs = -Q_base.copy()
        A = J.copy()
        for k in range(4):
            max_val = abs(A[k, k])
            max_row = k
            for r in range(k+1, 4):
                if abs(A[r, k]) > max_val:
                    max_val = abs(A[r, k])
                    max_row = r
            if max_row != k:
                for cc in range(4):
                    A[k, cc], A[max_row, cc] = A[max_row, cc], A[k, cc]
                rhs[k], rhs[max_row] = rhs[max_row], rhs[k]
            pivot = A[k, k]
            if abs(pivot) < 1e-20:
                continue
            for r in range(k+1, 4):
                factor = A[r, k] / pivot
                for cc in range(k+1, 4):
                    A[r, cc] -= factor * A[k, cc]
                rhs[r] -= factor * rhs[k]

        delta = np.zeros(4)
        for i in range(3, -1, -1):
            s = rhs[i]
            for j2 in range(i+1, 4):
                s -= A[i, j2] * delta[j2]
            if abs(A[i, i]) > 1e-20:
                delta[i] = s / A[i, i]

        phi = _clip(phi + delta[0], eps, 1.0 - eps)
        phi0 = _clip(phi0 + delta[1], eps, 1.0 - eps)
        psi = _clip(psi + delta[2], eps, 1.0 - eps)
        gamma = _clip(gamma + delta[3], -1e6, 1e6)

    return phi, phi0, psi, gamma


@njit(cache=True)
def _newton_step_felix_dynamic_c_exogenous(phi_old, phi0_old, psi_old,
                                            gamma_old,
                                            a11, b11, dt, Kp, eta,
                                            eta_phi, alpha, n_iter):
    eps = 1e-10
    eps_fd = 1e-8

    phi = _clip(phi_old, eps, 1.0 - eps)
    phi0 = _clip(phi0_old, eps, 1.0 - eps)
    psi = _clip(psi_old, eps, 1.0 - eps)
    gamma = _clip(gamma_old, -1e6, 1e6)

    for _ in range(n_iter):
        Q0, Q1, Q2, Q3 = _residual_felix_dynamic_c_exogenous(
            phi, phi0, psi, gamma,
            phi_old, phi0_old, psi_old,
            a11, b11, dt, Kp, eta, eta_phi, alpha
        )
        Q_base = np.array([Q0, Q1, Q2, Q3])

        # Numerical Jacobian
        J = np.zeros((4, 4))
        g = np.array([phi, phi0, psi, gamma])
        for j in range(4):
            gp = g.copy()
            gp[j] += eps_fd
            gp[0] = _clip(gp[0], eps, 1.0 - eps)
            gp[1] = _clip(gp[1], eps, 1.0 - eps)
            gp[2] = _clip(gp[2], eps, 1.0 - eps)
            gp[3] = _clip(gp[3], -1e6, 1e6)
            Qp0, Qp1, Qp2, Qp3 = _residual_felix_dynamic_c_exogenous(
                gp[0], gp[1], gp[2], gp[3],
                phi_old, phi0_old, psi_old,
                a11, b11, dt, Kp, eta, eta_phi, alpha
            )
            J[0, j] = (Qp0 - Q_base[0]) / eps_fd
            J[1, j] = (Qp1 - Q_base[1]) / eps_fd
            J[2, j] = (Qp2 - Q_base[2]) / eps_fd
            J[3, j] = (Qp3 - Q_base[3]) / eps_fd

        # Solve 4x4 with partial pivoting
        rhs = -Q_base.copy()
        A = J.copy()
        for k in range(4):
            max_val = abs(A[k, k])
            max_row = k
            for r in range(k + 1, 4):
                if abs(A[r, k]) > max_val:
                    max_val = abs(A[r, k])
                    max_row = r
            if max_row != k:
                for cc in range(4):
                    A[k, cc], A[max_row, cc] = A[max_row, cc], A[k, cc]
                rhs[k], rhs[max_row] = rhs[max_row], rhs[k]
            pivot = A[k, k]
            if abs(pivot) < 1e-20:
                continue
            for r in range(k + 1, 4):
                factor = A[r, k] / pivot
                for cc in range(k + 1, 4):
                    A[r, cc] -= factor * A[k, cc]
                rhs[r] -= factor * rhs[k]

        delta = np.zeros(4)
        for i in range(3, -1, -1):
            s = rhs[i]
            for j2 in range(i + 1, 4):
                s -= A[i, j2] * delta[j2]
            if abs(A[i, i]) > 1e-20:
                delta[i] = s / A[i, i]

        phi = _clip(phi + delta[0], eps, 1.0 - eps)
        phi0 = _clip(phi0 + delta[1], eps, 1.0 - eps)
        psi = _clip(psi + delta[2], eps, 1.0 - eps)
        gamma = _clip(gamma + delta[3], -1e6, 1e6)

    return phi, phi0, psi, gamma


@njit(cache=True)
def _newton_step_felix_dynamic_c_exogenous_scale(phi_old, phi0_old, psi_old,
                                                gamma_old,
                                                a11, b11, dt, Kp, eta,
                                                eta_phi, c_scale, alpha,
                                                n_iter):
    eps = 1e-10
    eps_fd = 1e-8

    phi = _clip(phi_old, eps, 1.0 - eps)
    phi0 = _clip(phi0_old, eps, 1.0 - eps)
    psi = _clip(psi_old, eps, 1.0 - eps)
    gamma = _clip(gamma_old, -1e6, 1e6)

    for _ in range(n_iter):
        Q0, Q1, Q2, Q3 = _residual_felix_dynamic_c_exogenous_scale(
            phi, phi0, psi, gamma,
            phi_old, phi0_old, psi_old,
            a11, b11, dt, Kp, eta, eta_phi, c_scale, alpha
        )
        Q_base = np.array([Q0, Q1, Q2, Q3])

        J = np.zeros((4, 4))
        g = np.array([phi, phi0, psi, gamma])
        for j in range(4):
            gp = g.copy()
            gp[j] += eps_fd
            gp[0] = _clip(gp[0], eps, 1.0 - eps)
            gp[1] = _clip(gp[1], eps, 1.0 - eps)
            gp[2] = _clip(gp[2], eps, 1.0 - eps)
            gp[3] = _clip(gp[3], -1e6, 1e6)
            Qp0, Qp1, Qp2, Qp3 = _residual_felix_dynamic_c_exogenous_scale(
                gp[0], gp[1], gp[2], gp[3],
                phi_old, phi0_old, psi_old,
                a11, b11, dt, Kp, eta, eta_phi, c_scale, alpha
            )
            J[0, j] = (Qp0 - Q_base[0]) / eps_fd
            J[1, j] = (Qp1 - Q_base[1]) / eps_fd
            J[2, j] = (Qp2 - Q_base[2]) / eps_fd
            J[3, j] = (Qp3 - Q_base[3]) / eps_fd

        rhs = -Q_base.copy()
        A = J.copy()
        for k in range(4):
            max_val = abs(A[k, k])
            max_row = k
            for r in range(k + 1, 4):
                if abs(A[r, k]) > max_val:
                    max_val = abs(A[r, k])
                    max_row = r
            if max_row != k:
                for cc in range(4):
                    A[k, cc], A[max_row, cc] = A[max_row, cc], A[k, cc]
                rhs[k], rhs[max_row] = rhs[max_row], rhs[k]
            pivot = A[k, k]
            if abs(pivot) < 1e-20:
                continue
            for r in range(k + 1, 4):
                factor = A[r, k] / pivot
                for cc in range(k + 1, 4):
                    A[r, cc] -= factor * A[k, cc]
                rhs[r] -= factor * rhs[k]

        delta = np.zeros(4)
        for i in range(3, -1, -1):
            s = rhs[i]
            for j2 in range(i + 1, 4):
                s -= A[i, j2] * delta[j2]
            if abs(A[i, i]) > 1e-20:
                delta[i] = s / A[i, i]

        phi = _clip(phi + delta[0], eps, 1.0 - eps)
        phi0 = _clip(phi0 + delta[1], eps, 1.0 - eps)
        psi = _clip(psi + delta[2], eps, 1.0 - eps)
        gamma = _clip(gamma + delta[3], -1e6, 1e6)

    return phi, phi0, psi, gamma


@njit(cache=True)
def _newton_step_felix_dynamic_c_state_substituted(phi_old, phi0_old, psi_old,
                                                     gamma_old,
                                                     a11, b11, dt, Kp,
                                                     eta, eta_phi, alpha,
                                                     n_iter):
    eps = 1e-10
    eps_fd = 1e-8

    phi = _clip(phi_old, eps, 1.0 - eps)
    phi0 = _clip(phi0_old, eps, 1.0 - eps)
    psi = _clip(psi_old, eps, 1.0 - eps)
    gamma = _clip(gamma_old, -1e6, 1e6)

    for _ in range(n_iter):
        Q0, Q1, Q2, Q3 = _residual_felix_dynamic_c_state_substituted(
            phi, phi0, psi, gamma,
            phi_old, phi0_old, psi_old,
            a11, b11, dt, Kp, eta, eta_phi, alpha
        )
        Q_base = np.array([Q0, Q1, Q2, Q3])

        # Numerical Jacobian
        J = np.zeros((4, 4))
        g = np.array([phi, phi0, psi, gamma])
        for j in range(4):
            gp = g.copy()
            gp[j] += eps_fd
            gp[0] = _clip(gp[0], eps, 1.0 - eps)
            gp[1] = _clip(gp[1], eps, 1.0 - eps)
            gp[2] = _clip(gp[2], eps, 1.0 - eps)
            gp[3] = _clip(gp[3], -1e6, 1e6)
            Qp0, Qp1, Qp2, Qp3 = _residual_felix_dynamic_c_state_substituted(
                gp[0], gp[1], gp[2], gp[3],
                phi_old, phi0_old, psi_old,
                a11, b11, dt, Kp, eta, eta_phi, alpha
            )
            J[0, j] = (Qp0 - Q_base[0]) / eps_fd
            J[1, j] = (Qp1 - Q_base[1]) / eps_fd
            J[2, j] = (Qp2 - Q_base[2]) / eps_fd
            J[3, j] = (Qp3 - Q_base[3]) / eps_fd

        # Solve 4x4 with partial pivoting
        rhs = -Q_base.copy()
        A = J.copy()
        for k in range(4):
            max_val = abs(A[k, k])
            max_row = k
            for r in range(k + 1, 4):
                if abs(A[r, k]) > max_val:
                    max_val = abs(A[r, k])
                    max_row = r
            if max_row != k:
                for cc in range(4):
                    A[k, cc], A[max_row, cc] = A[max_row, cc], A[k, cc]
                rhs[k], rhs[max_row] = rhs[max_row], rhs[k]
            pivot = A[k, k]
            if abs(pivot) < 1e-20:
                continue
            for r in range(k + 1, 4):
                factor = A[r, k] / pivot
                for cc in range(k + 1, 4):
                    A[r, cc] -= factor * A[k, cc]
                rhs[r] -= factor * rhs[k]

        delta = np.zeros(4)
        for i in range(3, -1, -1):
            s = rhs[i]
            for j2 in range(i + 1, 4):
                s -= A[i, j2] * delta[j2]
            if abs(A[i, i]) > 1e-20:
                delta[i] = s / A[i, i]

        phi = _clip(phi + delta[0], eps, 1.0 - eps)
        phi0 = _clip(phi0 + delta[1], eps, 1.0 - eps)
        psi = _clip(psi + delta[2], eps, 1.0 - eps)
        gamma = _clip(gamma + delta[3], -1e6, 1e6)

    return phi, phi0, psi, gamma


@njit(cache=True)
def simulate(a11, n_steps, dt=1e-4, phi_init=0.1,
             b11=0.0, Kp=1e-4, eta=1.0, eta_phi=1.0,
             c=100.0, alpha=0.0, n_newton=6):
    """
    Run monospecies Hamilton ODE (Felix formulation).
    Returns phi_hist, psi_hist, phibar_hist.
    """
    phi = phi_init
    phi0 = 1.0 - phi_init
    psi = 0.999
    gamma = 0.0

    phi_hist = np.empty(n_steps + 1)
    psi_hist = np.empty(n_steps + 1)
    phi_hist[0] = phi
    psi_hist[0] = psi

    for step in range(n_steps):
        phi, phi0, psi, gamma = _newton_step_felix(
            phi, phi0, psi, gamma,
            a11, b11, dt, Kp, eta, eta_phi, c, alpha, n_newton
        )
        phi_hist[step + 1] = phi
        psi_hist[step + 1] = psi

    return phi_hist, psi_hist, phi_hist * psi_hist


@njit(cache=True)
def simulate_felix_dynamic_c_exogenous(a11, n_steps, dt=1e-4, phi_init=0.1,
                                       b11=0.0, Kp=1e-4, eta=1.0,
                                       eta_phi=1.0, alpha=0.0,
                                       n_newton=6):
    """Monospecies simulation with c(t)=10*(1-phi*psi) (exogenous variant)."""
    phi = phi_init
    phi0 = 1.0 - phi_init
    psi = 0.999
    gamma = 0.0

    phi_hist = np.empty(n_steps + 1)
    psi_hist = np.empty(n_steps + 1)
    phi_hist[0] = phi
    psi_hist[0] = psi

    for step in range(n_steps):
        phi, phi0, psi, gamma = _newton_step_felix_dynamic_c_exogenous(
            phi, phi0, psi, gamma,
            a11, b11, dt, Kp, eta, eta_phi, alpha, n_newton
        )
        phi_hist[step + 1] = phi
        psi_hist[step + 1] = psi

    return phi_hist, psi_hist, phi_hist * psi_hist


@njit(cache=True)
def simulate_felix_dynamic_c_exogenous_scale(a11, n_steps, dt=1e-4, phi_init=0.1,
                                             b11=0.0, Kp=1e-4, eta=1.0,
                                             eta_phi=1.0, c_scale=10.0,
                                             alpha=0.0, n_newton=6):
    phi = phi_init
    phi0 = 1.0 - phi_init
    psi = 0.999
    gamma = 0.0

    phi_hist = np.empty(n_steps + 1)
    psi_hist = np.empty(n_steps + 1)
    phi_hist[0] = phi
    psi_hist[0] = psi

    for step in range(n_steps):
        phi, phi0, psi, gamma = _newton_step_felix_dynamic_c_exogenous_scale(
            phi, phi0, psi, gamma,
            a11, b11, dt, Kp, eta, eta_phi, c_scale, alpha, n_newton
        )
        phi_hist[step + 1] = phi
        psi_hist[step + 1] = psi

    return phi_hist, psi_hist, phi_hist * psi_hist


@njit(cache=True)
def simulate_felix_dynamic_c_exogenous_scale_sigma_series(a11_series, n_steps,
                                                         dt=1e-4, phi_init=0.1,
                                                         b11=0.0, Kp=1e-4,
                                                         eta=1.0, eta_phi=1.0,
                                                         c_scale=10.0,
                                                         alpha=0.0,
                                                         n_newton=6):
    phi = phi_init
    phi0 = 1.0 - phi_init
    psi = 0.999
    gamma = 0.0

    phi_hist = np.empty(n_steps + 1)
    psi_hist = np.empty(n_steps + 1)
    phi_hist[0] = phi
    psi_hist[0] = psi

    for step in range(n_steps):
        a11 = a11_series[step]
        phi, phi0, psi, gamma = _newton_step_felix_dynamic_c_exogenous_scale(
            phi, phi0, psi, gamma,
            a11, b11, dt, Kp, eta, eta_phi, c_scale, alpha, n_newton
        )
        phi_hist[step + 1] = phi
        psi_hist[step + 1] = psi

    return phi_hist, psi_hist, phi_hist * psi_hist


@njit(cache=True)
def simulate_felix_dynamic_c_state_substituted(a11, n_steps, dt=1e-4,
                                                  phi_init=0.1, b11=0.0,
                                                  Kp=1e-4, eta=1.0,
                                                  eta_phi=1.0, alpha=0.0,
                                                  n_newton=6):
    """Monospecies simulation with c(t)=10*(1-phi*psi) (state-substituted)."""
    phi = phi_init
    phi0 = 1.0 - phi_init
    psi = 0.999
    gamma = 0.0

    phi_hist = np.empty(n_steps + 1)
    psi_hist = np.empty(n_steps + 1)
    phi_hist[0] = phi
    psi_hist[0] = psi

    for step in range(n_steps):
        phi, phi0, psi, gamma = _newton_step_felix_dynamic_c_state_substituted(
            phi, phi0, psi, gamma,
            a11, b11, dt, Kp, eta, eta_phi, alpha, n_newton
        )
        phi_hist[step + 1] = phi
        psi_hist[step + 1] = psi

    return phi_hist, psi_hist, phi_hist * psi_hist


# ── Data loading ─────────────────────────────────────────────────────────────

def load_static_data(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['St']
    temps = [4, 8, 15, 20, 25, 35, 37, 40]
    data = {}
    for i, temp in enumerate(temps):
        col_start = i * 3
        times_all, cfu_all = [], []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
            t_cell, c_cell = row[col_start], row[col_start + 1]
            if t_cell.value is not None and c_cell.value is not None:
                times_all.append(float(t_cell.value))
                cfu_all.append(float(c_cell.value))
        unique_times = sorted(set(times_all))
        avg_cfu = [np.mean([cfu_all[j] for j in range(len(times_all)) if times_all[j] == t])
                   for t in unique_times]
        data[temp] = {'time': np.array(unique_times), 'cfu_mean': np.array(avg_cfu)}
    return data


def load_simulation_data(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['simulation']
    temps = [4, 8, 15, 20, 25, 35, 37, 40]
    sim_data = {}
    for i, temp in enumerate(temps):
        col_start = i * 11 + 1
        steps, phis, psis = [], [], []
        for r in range(3, ws.max_row + 1):
            sv = ws.cell(row=r, column=col_start).value
            pv = ws.cell(row=r, column=col_start + 1).value
            psv = ws.cell(row=r, column=col_start + 3).value
            if sv is None or pv is None: break
            steps.append(int(sv))
            phis.append(float(pv))
            psis.append(float(psv) if psv else 0.0)
        sim_data[temp] = {'step': np.array(steps), 'phi': np.array(phis), 'psi': np.array(psis)}
    return sim_data


# ── RMSE & Optimization ─────────────────────────────────────────────────────

def compute_rmse(a11, c_val, exp_time, exp_cfu, n_steps, time_to_step=10.0):
    """RMSE with optimal affine mapping φ̄(t) → log₁₀(CFU/mL)."""
    try:
        _, _, phibar = simulate(a11, n_steps, c=c_val)
    except:
        return 1e10
    idx = np.clip((exp_time * time_to_step).astype(int), 0, n_steps)
    sim_vals = phibar[idx]
    if np.std(sim_vals) < 1e-10:
        return 1e10
    X = np.column_stack([sim_vals, np.ones_like(sim_vals)])
    try:
        coeff, _, _, _ = np.linalg.lstsq(X, exp_cfu, rcond=None)
    except:
        return 1e10
    if coeff[0] < 0:
        return 1e10
    return np.sqrt(np.mean((X @ coeff - exp_cfu)**2))


def optimize_a11(exp_time, exp_cfu, n_steps, c=100.0,
                 a11_bounds=(0.1, 500.0)):
    """Find optimal a₁₁ for one temperature condition."""
    def obj(x):
        try:
            return compute_rmse(x[0], c, exp_time, exp_cfu, n_steps)
        except:
            return 1e10
    result = differential_evolution(obj, bounds=[a11_bounds],
                                    seed=42, maxiter=300, tol=1e-10, polish=True)
    return result.x[0], result.fun


# ── Main ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import os
    xlsx_path = os.path.join(os.path.dirname(__file__), 'raw data.xlsx')

    # Warmup Numba
    print("[Warmup] Compiling...")
    t0 = _time.time()
    _ = simulate(25.0, 100, c=100.0)
    print(f"  Done in {_time.time()-t0:.1f}s")

    # Step 1: Verify against Felix's simulation sheet
    print("\n[Step 1] Verify against Felix simulation (4°C, a₁₁=25, c=100)...")
    sim_data = load_simulation_data(xlsx_path)
    paper_a11 = {4: 25, 8: 50, 15: 50, 20: 70, 25: 70, 35: 110, 37: 110, 40: 50}

    for temp in [4, 8, 15, 20, 25, 35, 37, 40]:
        a11 = paper_a11[temp]
        sd = sim_data[temp]
        ns = sd['step'][-1]

        phi_ours, psi_ours, _ = simulate(float(a11), ns, c=100.0)

        print(f"\n  {temp}°C (a₁₁={a11}, {ns} steps):")
        check = [100, 500, 1000, min(2000, ns), ns]
        for s in check:
            idx_f = np.searchsorted(sd['step'], s)
            if idx_f < len(sd['step']) and sd['step'][idx_f] == s:
                phi_f = sd['phi'][idx_f]
                phi_o = phi_ours[s]
                print(f"    step {s:>5d}: Felix φ={phi_f:.6f}  Ours φ={phi_o:.6f}  Δ={phi_o-phi_f:.4e}")

    # Step 2: Optimize a₁₁ per temperature (c=100 fixed)
    print("\n\n[Step 2] Optimizing a₁₁ per temperature (c=100)...")
    data = load_static_data(xlsx_path)

    paper_rmse = {4: 0.0809, 8: 0.0476, 15: 0.1201, 20: 0.1580,
                  25: 0.1505, 35: 0.1690, 37: 0.2722, 40: 0.1150}
    max_steps = {4: 4000, 8: 2000, 15: 1000, 20: 500, 25: 300, 35: 200, 37: 200, 40: 450}

    print("=" * 115)
    print(f"{'Temp':>6s} | {'Paper a₁₁':>9s} {'Paper RMSE':>11s} | "
          f"{'Our a₁₁':>10s} {'Our RMSE':>10s} | {'Improve':>8s} {'Time':>6s}")
    print("=" * 115)

    for temp in [4, 8, 15, 20, 25, 35, 37, 40]:
        d = data[temp]
        ns = max_steps[temp]
        t0 = _time.time()

        a11_opt, rmse_opt = optimize_a11(d['time'], d['cfu_mean'], ns, c=100.0)

        elapsed = _time.time() - t0
        p_rmse = paper_rmse[temp]
        improve = (p_rmse - rmse_opt) / p_rmse * 100

        print(f"{temp:>4d}°C | {paper_a11[temp]:>9.1f} {p_rmse:>11.4f} | "
              f"{a11_opt:>10.4f} {rmse_opt:>10.6f} | "
              f"{improve:>+7.1f}% {elapsed:>5.1f}s")

    print("=" * 115)
