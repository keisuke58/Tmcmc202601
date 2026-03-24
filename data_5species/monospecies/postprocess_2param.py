#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
postprocess_2param.py — 8-panel visualization for 2-parameter TMCMC results.

theta = [sigma, phi_init]
Generates high-quality figure with MAP fit, 95% posterior predictive CI, and data.

Usage:
    python postprocess_2param.py \
        --indir _tmcmc_results_dyncc10_2param \
        --outpng figures/monospecies_8panel_dyncc10_2param.png \
        --n_pp_samples 500
"""

import os
import sys
import argparse
import numpy as np

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.lines
import matplotlib.patches

sys.path.insert(0, os.path.dirname(__file__))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'main'))

# Use JAX version (no numba dependency)
import os as _os
_os.environ.setdefault('XLA_PYTHON_CLIENT_PREALLOCATE', 'false')
import jax
import jax.numpy as jnp
jax.config.update('jax_enable_x64', True)
from hamilton_monospecies_jax import simulate_monospecies_phibar_hist_dynamic_c

TEMPS = [4, 8, 15, 20, 25, 35, 37, 40]
MAX_STEPS = {4: 4000, 8: 2000, 15: 1000, 20: 500,
             25: 300,  35: 200,  37: 200, 40: 450}
TIME_TO_STEP = 10.0

FELIX_SIGMA = {4: 4.25, 8: 10.0, 15: 25.0, 20: 50.0,
               25: 100.0, 35: 110.0, 37: 115.0, 40: 40.0}


def load_static_data(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['St']
    data = {}
    for i, temp in enumerate(TEMPS):
        col_start = i * 3
        times_rep, vals_rep = [], []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            t_val = row[col_start]
            c_val = row[col_start + 1]
            if t_val is not None and c_val is not None:
                times_rep.append(float(t_val))
                vals_rep.append(float(c_val))
        unique_times = sorted(set(times_rep))
        reps_by_time = {t: [] for t in unique_times}
        for t, v in zip(times_rep, vals_rep):
            reps_by_time[t].append(v)
        t_arr = np.array(unique_times, dtype=float)
        mean_arr = np.array([np.mean(reps_by_time[t]) for t in unique_times], dtype=float)
        std_arr = np.array([np.std(reps_by_time[t]) for t in unique_times], dtype=float)
        data[temp] = {
            'time': t_arr, 'mean': mean_arr, 'std': std_arr,
            'reps_by_time': reps_by_time,
        }
    return data


def make_batch_simulate(n_steps, c_scale=10.0, dt=1e-4, n_newton=6):
    """Create a vmapped batch simulator for given n_steps (JIT compiled once)."""
    from functools import partial

    @partial(jax.jit, static_argnums=())
    def single_sim(sigma, phi_init):
        return simulate_monospecies_phibar_hist_dynamic_c(
            sigma, n_steps, dt=dt,
            Kp=1e-4, eta=1.0, eta_phi=1.0,
            c_scale=float(c_scale), alpha=0.0, b=0.0,
            n_newton=n_newton, phi_init=phi_init,
        )

    batch_sim = jax.vmap(single_sim, in_axes=(0, 0))

    @partial(jax.jit)
    def batched(sigmas, phi_inits):
        return batch_sim(sigmas, phi_inits)

    return batched


def profiled_predictions_batch(phibar_batch, obs_step_idx, obs_mean):
    """Fit affine map for a batch of simulations. Returns (N, n_obs) predictions."""
    n_obs = len(obs_mean)
    sim_vals = phibar_batch[:, obs_step_idx]  # (N, n_obs)
    # vectorized least squares: per-sample
    obs_mean_jax = jnp.array(obs_mean)
    ones = jnp.ones((sim_vals.shape[1],))

    def fit_one(sv):
        X = jnp.column_stack([sv, ones])
        coeff, _, _, _ = jnp.linalg.lstsq(X, obs_mean_jax, rcond=None)
        return X @ coeff

    preds = jax.vmap(fit_one)(sim_vals)  # (N, n_obs)
    return preds


def profiled_predictions(phibar, obs_time, obs_mean, n_steps):
    """Fit affine map for a single simulation: cfu ≈ a*phibar + b."""
    step_idx = np.clip((np.array(obs_time) * TIME_TO_STEP).astype(int), 0, n_steps)
    sim_vals = np.array(phibar)[step_idx]
    X = np.column_stack([sim_vals, np.ones_like(sim_vals)])
    coeff, _, _, _ = np.linalg.lstsq(X, np.array(obs_mean), rcond=None)
    pred = X @ coeff
    return pred, coeff


def compute_rmse(pred, obs):
    return float(np.sqrt(np.mean((pred - np.array(obs)) ** 2)))


def posterior_predictive_ci(samples, obs_time, obs_mean, n_steps,
                             c_scale=10.0, dt=1e-4, n_newton=6,
                             batch_size=100):
    """
    samples: (N, 2) array of [sigma, phi_init] pairs
    Batched vmap evaluation — much faster than sequential.
    Returns (lo, hi) 95% CI at obs_time.
    """
    step_idx = np.clip((np.array(obs_time) * TIME_TO_STEP).astype(int), 0, n_steps)
    obs_mean_arr = np.array(obs_mean)
    batch_fn = make_batch_simulate(n_steps, c_scale=c_scale, dt=dt, n_newton=n_newton)

    all_preds = []
    for i in range(0, len(samples), batch_size):
        batch = samples[i:i + batch_size]
        sigmas = jnp.array(batch[:, 0], dtype=jnp.float64)
        phi_inits = jnp.array(batch[:, 1], dtype=jnp.float64)
        phibar_batch = np.asarray(batch_fn(sigmas, phi_inits))  # (B, n_steps+1)
        # fit affine per sample
        for j in range(len(batch)):
            sv = phibar_batch[j][step_idx]
            X = np.column_stack([sv, np.ones_like(sv)])
            coeff, _, _, _ = np.linalg.lstsq(X, obs_mean_arr, rcond=None)
            all_preds.append(X @ coeff)

    preds = np.array(all_preds)
    lo = np.percentile(preds, 2.5, axis=0)
    hi = np.percentile(preds, 97.5, axis=0)
    return lo, hi


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--indir', type=str,
                        default='_tmcmc_results_dyncc10_2param')
    parser.add_argument('--outpng', type=str,
                        default='figures/monospecies_8panel_dyncc10_2param.png')
    parser.add_argument('--c-scale', type=float, default=10.0)
    parser.add_argument('--dt', type=float, default=1e-4)
    parser.add_argument('--n-newton', type=int, default=6)
    parser.add_argument('--n_pp_samples', type=int, default=500)
    parser.add_argument('--seed', type=int, default=0)
    args = parser.parse_args()

    xlsx_path = os.path.join(os.path.dirname(__file__), 'raw data.xlsx')
    exp_data = load_static_data(xlsx_path)
    os.makedirs(os.path.dirname(os.path.abspath(args.outpng)), exist_ok=True)

    rng = np.random.default_rng(args.seed)

    fig = plt.figure(figsize=(20, 8))
    gs = gridspec.GridSpec(2, 4, figure=fig, hspace=0.55, wspace=0.32)

    summary_rows = []

    for idx_temp, temp in enumerate(TEMPS):
        ax = fig.add_subplot(gs[idx_temp // 4, idx_temp % 4])
        d = exp_data[temp]
        n_steps = MAX_STEPS[temp]

        npz_path = os.path.join(args.indir, f'samples_{temp}C.npz')
        npz = np.load(npz_path)
        samples = npz['samples']  # (N, 2)
        sigma_MAP = float(npz['sigma_MAP'])
        phi_init_MAP = float(npz['phi_init_MAP'])

        # MAP trajectory
        phibar_map = simulate_phibar(sigma_MAP, n_steps, c_scale=args.c_scale,
                                      dt=args.dt, n_newton=args.n_newton,
                                      phi_init=phi_init_MAP)
        pred_map_obs, coeff = profiled_predictions(phibar_map, d['time'], d['mean'], n_steps)
        rmse = compute_rmse(pred_map_obs, d['mean'])

        t_full = np.arange(n_steps + 1) / TIME_TO_STEP
        mapped_map_full = coeff[0] * phibar_map + coeff[1]
        ax.plot(t_full, mapped_map_full, '-', color='steelblue', lw=1.8, zorder=3, label='MAP fit')

        # Posterior predictive CI
        n_sel = min(args.n_pp_samples, len(samples))
        sel_idx = rng.choice(len(samples), size=n_sel, replace=False)
        lo, hi = posterior_predictive_ci(
            samples[sel_idx], d['time'], d['mean'], n_steps,
            c_scale=args.c_scale, dt=args.dt, n_newton=args.n_newton
        )
        ax.fill_between(d['time'], lo, hi, alpha=0.25, color='steelblue', label='95% CI')

        # Experimental data
        for t_pt, reps in d['reps_by_time'].items():
            for v in reps:
                ax.plot(t_pt, v, 'o', color='#e05a2b', ms=4, alpha=0.85, zorder=4,
                        markeredgewidth=0.4, markeredgecolor='white')
        ax.errorbar(d['time'], d['mean'], yerr=d['std'],
                    fmt='none', ecolor='#e05a2b', elinewidth=1.0,
                    capsize=2, zorder=5, alpha=0.7)

        title = (f'({list("abcdefgh")[idx_temp]}) {temp}°C  '
                 f'σ={sigma_MAP:.2f} φ₀={phi_init_MAP:.3f}  RMSE={rmse:.3f}')
        ax.set_title(title, fontsize=9, pad=5)
        ax.set_xlabel('Time (min)', fontsize=9)
        ax.set_ylabel('log₁₀(CFU/mL)', fontsize=9)
        ax.tick_params(labelsize=8)
        ax.spines[['top', 'right']].set_visible(False)

        if idx_temp == 0:
            ax.legend(fontsize=7, framealpha=0.7,
                      handles=[
                          matplotlib.lines.Line2D([], [], color='steelblue', lw=1.8, label='MAP fit'),
                          matplotlib.patches.Patch(color='steelblue', alpha=0.25, label='95% CI'),
                          matplotlib.lines.Line2D([], [], marker='o', color='w',
                                                  markerfacecolor='#e05a2b', ms=5, label='Experiment'),
                      ])

        summary_rows.append({
            'temp': temp,
            'felix_sigma': FELIX_SIGMA[temp],
            'sigma_MAP_1D': '(see 1D run)',
            'sigma_MAP_2D': sigma_MAP,
            'phi_init_MAP': phi_init_MAP,
            'RMSE_2D': rmse,
        })
        print(f"  {temp}°C: σ={sigma_MAP:.3f}, φ₀={phi_init_MAP:.4f}, RMSE={rmse:.4f}")

    fig.suptitle('Monospecies TMCMC — 2-param (σ, φ₀) fit  [dyncc10, RW, 4000p]',
                 fontsize=12, y=1.01)
    fig.savefig(args.outpng, dpi=300, bbox_inches='tight')
    plt.close(fig)
    print(f"\nSaved: {args.outpng}")

    # Compare with 1D
    indir_1d = args.indir.replace('_2param', '_rw4000p')
    if os.path.isdir(indir_1d):
        print(f"\n{'='*70}")
        print(f"{'Temp':>5s} | {'Felix σ':>8s} | {'1D σ':>8s} | {'2D σ':>8s} | {'2D φ₀':>8s} | {'RMSE_2D':>8s}")
        print('-' * 70)
        for row in summary_rows:
            temp = row['temp']
            npz1d = os.path.join(indir_1d, f'samples_{temp}C.npz')
            if os.path.exists(npz1d):
                d1 = np.load(npz1d)
                sigma1d = float(d1['sigma_MAP'])
            else:
                sigma1d = float('nan')
            print(f"{temp:>4d}°C | {row['felix_sigma']:>8.2f} | {sigma1d:>8.3f} | "
                  f"{row['sigma_MAP_2D']:>8.3f} | {row['phi_init_MAP']:>8.4f} | "
                  f"{row['RMSE_2D']:>8.4f}")


if __name__ == '__main__':
    main()
