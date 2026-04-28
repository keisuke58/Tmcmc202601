#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Reproduce `raw data.xlsx` -> `calc_error` mapping and RMSE calculation.

This script matches the Excel formulas used in the `calc_error` sheet:

- mapping = slope * result + offset
- abs error = mapping - avg Con
- abs error sq = (abs error)^2
- RMSE = sqrt( sum(abs error sq) / sum(mapping) )

It then compares the computed RMSE against the sheet's RMSE cell for each
temperature block.
"""

from __future__ import annotations

import argparse
import csv
import logging
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl


logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class BlockSpec:
    temp_c: int
    start_col: int  # calc_error block start column index (1-based)

    @property
    def rmse_col(self) -> int:
        # Excel uses: paper RMSE cell = sqrt(I15), and for each block
        # the RMSE cell is located at start_col + 9 on row 15.
        return self.start_col + 9

    @property
    def time_col(self) -> int:
        return self.start_col

    @property
    def avgcon_col(self) -> int:
        return self.start_col + 1

    @property
    def result_col(self) -> int:
        return self.start_col + 3

    @property
    def mapping_col(self) -> int:
        return self.start_col + 4

    @property
    def abs_error_sq_col(self) -> int:
        return self.start_col + 7


def _num(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        if math.isnan(v):
            return None
        return float(v)
    return None


def read_mapping_constants(ws) -> Tuple[float, float]:
    """
    Read slope and offset used in `calc_error` -> `mapping` for all blocks.

    In the sheet, slope/offset are derived from:
      ymin sim = B18, ymax sim = C18
      ymin exp = D18, ymax exp = E18
    and:
      slope  = (ymax exp - ymin exp) / (ymax sim - ymin sim)
      offset = ymin exp - slope * ymin sim
    """
    ymin_sim = _num(ws["B18"].value)
    ymax_sim = _num(ws["C18"].value)
    ymin_exp = _num(ws["D18"].value)
    ymax_exp = _num(ws["E18"].value)
    if None in (ymin_sim, ymax_sim, ymin_exp, ymax_exp):
        raise RuntimeError(
            "Failed to read mapping constants from calc_error sheet "
            "(cells B18..E18 must be numeric)."
        )

    denom = ymax_sim - ymin_sim
    if abs(denom) < 1e-15:
        raise RuntimeError("Invalid mapping constants: ymax_sim == ymin_sim.")

    slope = (ymax_exp - ymin_exp) / denom
    offset = ymin_exp - slope * ymin_sim
    return slope, offset


def extract_block_series(
    ws,
    spec: BlockSpec,
    row_start: int = 4,
    row_end: int | None = None,
) -> Tuple[List[float], List[float], List[float]]:
    """
    Extract (avg_con, result, mapping) series from a block.

    The sheet stores a contiguous set of observation rows starting at row 4.
    We stop when `time_col` becomes empty for the first time after having
    collected at least one numeric point.
    """
    if row_end is None:
        row_end = ws.max_row

    avgcon: List[float] = []
    result: List[float] = []
    mapping: List[float] = []

    have_any = False
    for r in range(row_start, row_end + 1):
        t = _num(ws.cell(row=r, column=spec.time_col).value)
        if t is None:
            if have_any:
                break
            continue

        avgv = _num(ws.cell(row=r, column=spec.avgcon_col).value)
        resv = _num(ws.cell(row=r, column=spec.result_col).value)
        mapv = _num(ws.cell(row=r, column=spec.mapping_col).value)
        if None in (avgv, resv, mapv):
            continue

        have_any = True
        avgcon.append(float(avgv))
        result.append(float(resv))
        mapping.append(float(mapv))

    if not have_any:
        raise RuntimeError(f"No data rows found for {spec.temp_c}°C block.")

    return avgcon, result, mapping


def compute_rmse_from_series(
    avgcon: List[float],
    result: List[float],
    slope: float,
    offset: float,
) -> Tuple[float, float]:
    """
    Compute mapping and RMSE matching Excel formulas.

    Excel mapping:
      mapping = slope * result + offset
    abs error sq = (mapping - avgcon)^2
    RMSE = sqrt( sum(abs error sq) / sum(mapping) )
    """
    if len(avgcon) != len(result):
        raise ValueError("avgcon and result series must have same length.")

    mapping = [slope * x + offset for x in result]
    abs_err_sq = [(m - y) ** 2 for (m, y) in zip(mapping, avgcon)]

    sum_mapping = sum(mapping)
    if abs(sum_mapping) < 1e-15:
        return float("nan"), float("nan")

    sum_abs_err_sq = sum(abs_err_sq)
    rmse = math.sqrt(sum_abs_err_sq / sum_mapping)
    return rmse, sum_abs_err_sq


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--xlsx",
        default=str(Path(__file__).resolve().parent / "raw data.xlsx"),
        help="Path to raw data.xlsx",
    )
    parser.add_argument(
        "--out-csv",
        default=str(Path(__file__).resolve().parent / "calc_error_mapping_rmse_reproduce.csv"),
        help="Output CSV path",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable debug logging",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise FileNotFoundError(str(xlsx_path))

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["calc_error"]

    slope, offset = read_mapping_constants(ws)
    logger.info("mapping: slope=%g offset=%g", slope, offset)

    blocks = [
        BlockSpec(temp_c=4, start_col=1),
        BlockSpec(temp_c=8, start_col=12),
        BlockSpec(temp_c=15, start_col=23),
        BlockSpec(temp_c=20, start_col=34),
        BlockSpec(temp_c=25, start_col=45),
        BlockSpec(temp_c=35, start_col=56),
        BlockSpec(temp_c=37, start_col=67),
        BlockSpec(temp_c=40, start_col=78),
    ]

    rows: List[Dict[str, float | int | str]] = []
    for spec in blocks:
        avgcon, result, mapping_sheet = extract_block_series(ws, spec)

        rmse_calc, _ = compute_rmse_from_series(
            avgcon=avgcon,
            result=result,
            slope=slope,
            offset=offset,
        )

        rmse_excel = _num(ws.cell(row=15, column=spec.rmse_col).value)
        if rmse_excel is None:
            raise RuntimeError(f"Failed to read Excel RMSE at row15 col {spec.rmse_col} for {spec.temp_c}°C")

        rows.append(
            {
                "temp_c": spec.temp_c,
                "rmse_excel": rmse_excel,
                "rmse_calc": rmse_calc,
                "abs_diff": abs(rmse_calc - rmse_excel),
                "n_obs": len(avgcon),
            }
        )

        logger.info(
            "%d°C rmse_excel=%g rmse_calc=%g diff=%g n=%d",
            spec.temp_c,
            rmse_excel,
            rmse_calc,
            abs(rmse_calc - rmse_excel),
            len(avgcon),
        )

        # Optional: check mapping equality with sheet mapping values.
        # If this fails, the sheet may not be recalculated or columns are misaligned.
        mapping_calc = [slope * x + offset for x in result]
        max_map_diff = max(abs(a - b) for a, b in zip(mapping_calc, mapping_sheet))
        logger.debug("%d°C max mapping diff vs sheet: %g", spec.temp_c, max_map_diff)

    out_csv = Path(args.out_csv)
    out_csv.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = ["temp_c", "rmse_excel", "rmse_calc", "abs_diff", "n_obs"]
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    logger.info("Wrote: %s", str(out_csv))


if __name__ == "__main__":
    main()

