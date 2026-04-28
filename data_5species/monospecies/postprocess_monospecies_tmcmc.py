#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import json
import argparse
import csv
from pathlib import Path

import numpy as np

from hamilton_monospecies import (
    simulate,
    simulate_felix_dynamic_c_exogenous_scale,
    simulate_felix_dynamic_c_exogenous_scale_sigma_series,
)

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec


TEMPS = [4, 8, 15, 20, 25, 35, 37, 40]
MAX_STEPS = {4: 4000, 8: 2000, 15: 1000, 20: 500,
             25: 300,  35: 200,  37: 200, 40: 450}
TIME_TO_STEP = 10.0

PAPER_SIGMA = {4: 25, 8: 50, 15: 50, 20: 70, 25: 70, 35: 110, 37: 110, 40: 50}

FELIX_SIGMA = {4: 4.25, 8: 10.0, 15: 25.0, 20: 50.0,
               25: 100.0, 35: 110.0, 37: 115.0, 40: 40.0}


def load_static_data_with_reps(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['St']
    data = {}
    for i, temp in enumerate(TEMPS):
        col_start = i * 3
        times_rep, vals_rep = [], []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            t_val = row[col_start]
            c_val = row[col_start + 1]
            if t_val is not None and c_val is not None:
                times_rep.append(float(t_val))
                vals_rep.append(float(c_val))
        unique_times = sorted(set(times_rep))
        reps_by_time = {t: [] for t in unique_times}
        for t, v in zip(times_rep, vals_rep):
            reps_by_time[t].append(v)
        t_arr = np.array(unique_times, dtype=float)
        mean_arr = np.array([np.mean(reps_by_time[t]) for t in unique_times], dtype=float)
        std_arr = np.array([np.std(reps_by_time[t]) for t in unique_times], dtype=float)
        data[temp] = {
            'time': t_arr,
            'mean': mean_arr,
            'std': std_arr,
            'reps_by_time': reps_by_time,
        }
    return data


def load_dynamic_data_with_reps(xlsx_path, sheet='Dy'):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet]

    obs_time, obs_reps = [], []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        t, r1, r2, r3 = row[0], row[1], row[2], row[3]
        if t is None:
            continue
        reps = [v for v in (r1, r2, r3) if v is not None]
        if not reps:
            continue
        obs_time.append(float(t))
        obs_reps.append([float(v) for v in reps])

    unique_times = sorted(set(obs_time))
    reps_by_time = {t: [] for t in unique_times}
    for t, reps in zip(obs_time, obs_reps):
        reps_by_time[float(t)].extend(reps)

    t_arr = np.array(unique_times, dtype=float)
    mean_arr = np.array([np.mean(reps_by_time[t]) for t in unique_times], dtype=float)
    std_arr = np.array([np.std(reps_by_time[t]) for t in unique_times], dtype=float)

    temp_time, temp_val = [], []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        t, temp = row[5], row[6]
        if t is None or temp is None:
            continue
        temp_time.append(float(t))
        temp_val.append(float(temp))

    return {
        'time': t_arr,
        'mean': mean_arr,
        'std': std_arr,
        'reps_by_time': reps_by_time,
        'temp_time': np.array(temp_time, dtype=float),
        'temp_C': np.array(temp_val, dtype=float),
    }


def affine_fit(x, y):
    X = np.column_stack([x, np.ones_like(x)])
    coeff, _, _, _ = np.linalg.lstsq(X, y, rcond=None)
    return coeff


def profiled_predictions(phibar, obs_time, obs_mean, n_steps):
    idx = np.clip((obs_time * TIME_TO_STEP).astype(int), 0, n_steps)
    sim_at_obs = phibar[idx]
    coeff = affine_fit(sim_at_obs, obs_mean)
    X = np.column_stack([sim_at_obs, np.ones_like(sim_at_obs)])
    pred = X @ coeff
    return pred, coeff


def compute_metrics_at_obs(pred, obs_mean, obs_std, noise_floor):
    resid = pred - obs_mean
    rmse = float(np.sqrt(np.mean(resid**2)))
    mae = float(np.mean(np.abs(resid)))
    denom = float(np.sum((obs_mean - np.mean(obs_mean))**2))
    r2 = float(1.0 - np.sum(resid**2) / denom) if denom > 1e-16 else float('nan')
    sigma_obs = np.maximum(obs_std, noise_floor)
    z = resid / sigma_obs
    nll = float(0.5 * np.sum(z**2) + np.sum(np.log(sigma_obs)))
    logL = float(-nll)
    wrmse = float(np.sqrt(np.mean(z**2)))
    return {
        'rmse_mean': rmse,
        'mae_mean': mae,
        'r2_mean': r2,
        'logL_map': logL,
        'nll_map': nll,
        'wrmse_mean': wrmse,
    }


def simulate_phibar(sigma, n_steps, dynamic_c, c_scale, dt, n_newton):
    if dynamic_c:
        _, _, phibar = simulate_felix_dynamic_c_exogenous_scale(
            float(sigma), n_steps, dt=float(dt),
            Kp=float(1e-4), eta=float(1.0), eta_phi=float(1.0),
            c_scale=float(c_scale), alpha=float(0.0),
            n_newton=int(n_newton),
        )
        return np.asarray(phibar, dtype=float)
    _, _, phibar = simulate(
        float(sigma), n_steps, dt=float(dt),
        Kp=float(1e-4), eta=float(1.0), eta_phi=float(1.0),
        c=float(100.0), alpha=float(0.0),
        n_newton=int(n_newton),
    )
    return np.asarray(phibar, dtype=float)


def posterior_predictive_ci_obs(sigmas, obs_time, obs_mean, n_steps, dynamic_c, c_scale, dt, n_newton):
    sigmas = np.asarray(sigmas, dtype=float)
    obs_time = np.asarray(obs_time, dtype=float)
    obs_mean = np.asarray(obs_mean, dtype=float)
    idx = np.clip((obs_time * TIME_TO_STEP).astype(int), 0, n_steps)

    pred = np.empty((len(sigmas), len(obs_time)), dtype=float)
    for i, s in enumerate(sigmas):
        phibar = simulate_phibar(s, n_steps, dynamic_c, c_scale, dt, n_newton)
        x = phibar[idx]
        coeff = affine_fit(x, obs_mean)
        pred[i, :] = coeff[0] * x + coeff[1]

    lo = np.percentile(pred, 2.5, axis=0)
    hi = np.percentile(pred, 97.5, axis=0)
    return lo, hi, pred


def coverage_from_ci(reps_by_time, obs_time, lo, hi):
    total = 0
    inside = 0
    for t, l, h in zip(obs_time, lo, hi):
        reps = reps_by_time.get(float(t), [])
        for v in reps:
            total += 1
            if l <= v <= h:
                inside += 1
    return float(inside / total) if total > 0 else float('nan')


def interp_sigma_from_knots(temps_C, sigmas, T_query):
    temps_C = np.asarray(temps_C, dtype=float)
    sigmas = np.asarray(sigmas, dtype=float)
    T_query = np.asarray(T_query, dtype=float)
    order = np.argsort(temps_C)
    temps_C = temps_C[order]
    sigmas = sigmas[order]
    return np.interp(T_query, temps_C, sigmas, left=sigmas[0], right=sigmas[-1])


def fit_dynamic_scale_k(dy, sigma_knots, dt, n_newton, dynamic_c, c_scale, time_to_step_dy=10.0, noise_floor=0.1):
    t_obs = dy['time']
    y_obs = dy['mean']
    temp_time = dy['temp_time']
    temp_C = dy['temp_C']

    t_max = float(max(np.max(t_obs), np.max(temp_time)))
    n_steps = int(np.ceil(t_max * time_to_step_dy))
    t_step = np.arange(n_steps + 1, dtype=float) / float(time_to_step_dy)
    T_step = np.interp(t_step, temp_time, temp_C)

    temps = np.array([k for k in sigma_knots.keys()], dtype=float)
    sigmas = np.array([sigma_knots[k] for k in temps], dtype=float)
    sigma_base = interp_sigma_from_knots(temps, sigmas, T_step[:-1])

    idx_obs = np.clip((t_obs * time_to_step_dy).astype(int), 0, n_steps)

    def eval_k(k):
        sigma_series = np.clip(k * sigma_base, 1.0, 500.0)
        if dynamic_c:
            _, _, phibar = simulate_felix_dynamic_c_exogenous_scale_sigma_series(
                np.asarray(sigma_series, dtype=float),
                n_steps,
                dt=float(dt),
                Kp=float(1e-4),
                eta=float(1.0),
                eta_phi=float(1.0),
                c_scale=float(c_scale),
                alpha=float(0.0),
                n_newton=int(n_newton),
            )
        else:
            _, _, phibar = simulate(
                float(np.mean(sigma_series)),
                n_steps,
                dt=float(dt),
                Kp=float(1e-4),
                eta=float(1.0),
                eta_phi=float(1.0),
                c=float(100.0),
                alpha=float(0.0),
                n_newton=int(n_newton),
            )
        phibar = np.asarray(phibar, dtype=float)
        x = phibar[idx_obs]
        coeff = affine_fit(x, y_obs)
        pred = coeff[0] * x + coeff[1]
        m = compute_metrics_at_obs(pred, y_obs, dy['std'], noise_floor)
        return m['rmse_mean'], coeff, pred, phibar, T_step

    ks = np.linspace(0.2, 3.0, 29)
    best = None
    for k in ks:
        rmse, coeff, pred, phibar, T_step = eval_k(float(k))
        if best is None or rmse < best[0]:
            best = (rmse, float(k), coeff, pred, phibar, T_step)

    rmse0, k0, coeff0, pred0, phibar0, T_step0 = best
    ks2 = np.linspace(max(0.05, k0 - 0.3), k0 + 0.3, 31)
    for k in ks2:
        rmse, coeff, pred, phibar, T_step = eval_k(float(k))
        if rmse < rmse0:
            rmse0, k0, coeff0, pred0, phibar0, T_step0 = rmse, float(k), coeff, pred, phibar, T_step

    return {
        'k_best': float(k0),
        'rmse_best': float(rmse0),
        'affine_a': float(coeff0[0]),
        'affine_b': float(coeff0[1]),
        't_step': np.arange(len(phibar0), dtype=float) / float(time_to_step_dy),
        'temp_step_C': np.array(T_step0, dtype=float),
        'pred_at_obs': np.array(pred0, dtype=float),
        't_obs': np.array(t_obs, dtype=float),
        'y_obs_mean': np.array(y_obs, dtype=float),
        'y_obs_std': np.array(dy['std'], dtype=float),
    }


def plot_dynamic_overview(out_png, dy, fit=None):
    fig, (axT, axY) = plt.subplots(2, 1, figsize=(10, 6), sharex=True, gridspec_kw={'height_ratios': [1, 2]})

    axT.plot(dy['temp_time'], dy['temp_C'], color='black', lw=1.0)
    axT.set_ylabel('Temperature (°C)')
    axT.spines[['top', 'right']].set_visible(False)

    for t_pt, reps in dy['reps_by_time'].items():
        for v in reps:
            axY.plot(t_pt, v, 'o', color='#e05a2b', ms=4, alpha=0.85, markeredgewidth=0.4, markeredgecolor='white')
    axY.errorbar(dy['time'], dy['mean'], yerr=dy['std'], fmt='none', ecolor='#e05a2b', elinewidth=1.0, capsize=2, alpha=0.7)

    if fit is not None:
        axY.plot(fit['t_obs'], fit['pred_at_obs'], '-', color='steelblue', lw=2.0, label=f"fit (k={fit['k_best']:.2f}, RMSE={fit['rmse_best']:.2f})")
        axY.legend(framealpha=0.7, fontsize=9)

    axY.set_xlabel('Time (h)')
    axY.set_ylabel('log₁₀(CFU/mL)')
    axY.spines[['top', 'right']].set_visible(False)

    fig.tight_layout()
    fig.savefig(out_png, dpi=150, bbox_inches='tight')
    plt.close(fig)


def load_sigma_knots_from_dir(sigma_knots_dir):
    sigma_knots_dir = str(sigma_knots_dir)
    knots = {}
    for temp in TEMPS:
        path = os.path.join(sigma_knots_dir, f'samples_{temp}C.npz')
        if not os.path.exists(path):
            raise FileNotFoundError(path)
        z = np.load(path)
        knots[float(temp)] = float(z['sigma_MAP'])
    return knots


def dy_tmcmc_paths(dy_dir):
    dy_dir = str(dy_dir)
    return {
        'dir': dy_dir,
        'csv': os.path.join(dy_dir, 'tmcmc_summary_Dy_k.csv'),
        'json': os.path.join(dy_dir, 'tmcmc_summary_Dy_k.json'),
        'npz': os.path.join(dy_dir, 'samples_Dy_k.npz'),
    }


def load_dy_tmcmc_outputs(dy_dir):
    p = dy_tmcmc_paths(dy_dir)
    if not os.path.exists(p['json']):
        raise FileNotFoundError(p['json'])
    if not os.path.exists(p['npz']):
        raise FileNotFoundError(p['npz'])

    with open(p['json'], 'r') as f:
        meta = json.load(f)

    z = np.load(p['npz'])
    samples = np.asarray(z['samples'], dtype=float).reshape(-1)
    k_map = float(z['k_MAP']) if 'k_MAP' in z.files else float(np.median(samples))
    time_to_step = float(z['time_to_step']) if 'time_to_step' in z.files else float(meta.get('dy_time_to_step', TIME_TO_STEP))

    return {
        **p,
        'meta': meta,
        'samples_k': samples,
        'k_map': k_map,
        'time_to_step': time_to_step,
    }


def dy_build_sigma_base_from_knots(dy, sigma_knots, time_to_step):
    t_max = float(max(np.max(dy['time']), np.max(dy['temp_time'])))
    n_steps = int(np.ceil(t_max * time_to_step))
    t_grid = np.arange(n_steps + 1, dtype=float) / float(time_to_step)
    T_grid = np.interp(t_grid, dy['temp_time'], dy['temp_C'])
    temps = np.array(sorted(sigma_knots.keys()), dtype=float)
    sigmas = np.array([sigma_knots[t] for t in temps], dtype=float)
    sigma_base = np.interp(T_grid[:-1], temps, sigmas, left=sigmas[0], right=sigmas[-1])
    return n_steps, t_grid, T_grid, sigma_base


def dy_simulate_phibar_from_sigma_series(sigma_series, n_steps, dt, n_newton, c_scale):
    _, _, phibar = simulate_felix_dynamic_c_exogenous_scale_sigma_series(
        np.asarray(sigma_series, dtype=float),
        n_steps,
        dt=float(dt),
        Kp=float(1e-4),
        eta=float(1.0),
        eta_phi=float(1.0),
        c_scale=float(c_scale),
        alpha=float(0.0),
        n_newton=int(n_newton),
    )
    return np.asarray(phibar, dtype=float)


def plot_dy_tmcmc_fig1_like(out_base, dy, dy_tmcmc, sigma_knots, n_pp_samples=200, seed=0):
    plt.rcParams.update({
        'font.size': 10,
        'axes.labelsize': 10,
        'axes.titlesize': 11,
        'legend.fontsize': 9,
        'xtick.labelsize': 9,
        'ytick.labelsize': 9,
        'axes.linewidth': 0.8,
        'lines.linewidth': 1.8,
        'savefig.dpi': 600,
        'pdf.fonttype': 42,
        'ps.fonttype': 42,
    })

    meta = dy_tmcmc['meta']
    dt = float(meta.get('dt', 1e-4))
    n_newton = int(meta.get('n_newton', 6))
    c_scale = float(meta.get('c_scale', 10.0))
    time_to_step = float(dy_tmcmc['time_to_step'])

    n_steps, t_grid, T_grid, sigma_base = dy_build_sigma_base_from_knots(dy, sigma_knots, time_to_step)
    idx_obs = np.clip((dy['time'] * time_to_step).astype(int), 0, n_steps)

    k_map = float(dy_tmcmc['k_map'])
    sigma_series_map = np.clip(k_map * sigma_base, 1.0, 500.0)
    phibar_map = dy_simulate_phibar_from_sigma_series(sigma_series_map, n_steps, dt, n_newton, c_scale)
    x_obs = phibar_map[idx_obs]
    coeff_map = affine_fit(x_obs, dy['mean'])
    y_map_full = coeff_map[0] * phibar_map + coeff_map[1]

    rng = np.random.default_rng(seed)
    samples_k = np.asarray(dy_tmcmc['samples_k'], dtype=float)
    sel = rng.choice(len(samples_k), size=min(int(n_pp_samples), len(samples_k)), replace=False)
    k_sel = samples_k[sel]

    preds_full = np.empty((len(k_sel), n_steps + 1), dtype=np.float32)
    for i, k in enumerate(k_sel):
        sigma_series = np.clip(float(k) * sigma_base, 1.0, 500.0)
        phibar = dy_simulate_phibar_from_sigma_series(sigma_series, n_steps, dt, n_newton, c_scale)
        x_obs_i = phibar[idx_obs]
        coeff = affine_fit(x_obs_i, dy['mean'])
        preds_full[i, :] = (coeff[0] * phibar + coeff[1]).astype(np.float32)

    lo_full = np.percentile(preds_full, 2.5, axis=0)
    hi_full = np.percentile(preds_full, 97.5, axis=0)

    fig, (axT, axY) = plt.subplots(
        2, 1, figsize=(6.8, 4.8), sharex=True,
        gridspec_kw={'height_ratios': [1.0, 2.2]},
        constrained_layout=True,
    )

    axT.plot(t_grid, T_grid, color='black', lw=1.2)
    axT.set_ylabel('Temperature (°C)')
    axT.spines[['top', 'right']].set_visible(False)

    axY.fill_between(t_grid, lo_full, hi_full, color='steelblue', alpha=0.22, linewidth=0.0, label='95% CI')
    axY.plot(t_grid, y_map_full, '-', color='steelblue', lw=2.2, label=f"MAP (k={k_map:.3f})")

    for t_pt, reps in dy['reps_by_time'].items():
        for v in reps:
            axY.plot(t_pt, v, 'o', color='#e05a2b', ms=4.2, alpha=0.9, markeredgewidth=0.45, markeredgecolor='white')
    axY.errorbar(dy['time'], dy['mean'], yerr=dy['std'], fmt='none', ecolor='#e05a2b', elinewidth=1.0, capsize=2.5, alpha=0.75)

    axY.set_xlabel('Time (h)')
    axY.set_ylabel('log₁₀(CFU/mL)')
    axY.spines[['top', 'right']].set_visible(False)
    axY.legend(framealpha=0.85, loc='lower right')

    fig.savefig(out_base + '.pdf', bbox_inches='tight')
    fig.savefig(out_base + '.png', bbox_inches='tight')
    plt.close(fig)


def plot_dy_k_posterior(out_base, dy_tmcmc):
    plt.rcParams.update({
        'font.size': 10,
        'axes.labelsize': 10,
        'axes.titlesize': 11,
        'savefig.dpi': 600,
        'pdf.fonttype': 42,
        'ps.fonttype': 42,
    })

    samples = np.asarray(dy_tmcmc['samples_k'], dtype=float)
    k_map = float(dy_tmcmc['k_map'])

    fig, ax = plt.subplots(1, 1, figsize=(5.2, 2.8), constrained_layout=True)
    ax.hist(samples, bins=40, color='steelblue', alpha=0.85, density=True)
    ax.axvline(k_map, color='black', lw=1.6, label=f"MAP k={k_map:.3f}")
    ax.set_xlabel('k')
    ax.set_ylabel('Density')
    ax.spines[['top', 'right']].set_visible(False)
    ax.legend(framealpha=0.85)
    fig.savefig(out_base + '.pdf', bbox_inches='tight')
    fig.savefig(out_base + '.png', bbox_inches='tight')
    plt.close(fig)


def evaluate_dy_tmcmc_metrics(dy, dy_tmcmc, sigma_knots, n_pp_samples=500, seed=0, noise_floor=0.1):
    meta = dy_tmcmc['meta']
    dt = float(meta.get('dt', 1e-4))
    n_newton = int(meta.get('n_newton', 6))
    c_scale = float(meta.get('c_scale', 10.0))
    time_to_step = float(dy_tmcmc['time_to_step'])

    n_steps, t_grid, T_grid, sigma_base = dy_build_sigma_base_from_knots(dy, sigma_knots, time_to_step)
    idx_obs = np.clip((dy['time'] * time_to_step).astype(int), 0, n_steps)

    k_map = float(dy_tmcmc['k_map'])
    sigma_series_map = np.clip(k_map * sigma_base, 1.0, 500.0)
    phibar_map = dy_simulate_phibar_from_sigma_series(sigma_series_map, n_steps, dt, n_newton, c_scale)
    x_obs_map = phibar_map[idx_obs]
    coeff_map = affine_fit(x_obs_map, dy['mean'])
    pred_map_obs = coeff_map[0] * x_obs_map + coeff_map[1]

    m = compute_metrics_at_obs(pred_map_obs, dy['mean'], dy['std'], noise_floor)

    rng = np.random.default_rng(seed)
    samples_k = np.asarray(dy_tmcmc['samples_k'], dtype=float)
    sel = rng.choice(len(samples_k), size=min(int(n_pp_samples), len(samples_k)), replace=False)
    k_sel = samples_k[sel]

    pred_at_obs = np.empty((len(k_sel), len(dy['time'])), dtype=np.float32)
    for i, k in enumerate(k_sel):
        sigma_series = np.clip(float(k) * sigma_base, 1.0, 500.0)
        phibar = dy_simulate_phibar_from_sigma_series(sigma_series, n_steps, dt, n_newton, c_scale)
        x_obs = phibar[idx_obs]
        coeff = affine_fit(x_obs, dy['mean'])
        pred_at_obs[i, :] = (coeff[0] * x_obs + coeff[1]).astype(np.float32)

    lo = np.percentile(pred_at_obs, 2.5, axis=0)
    hi = np.percentile(pred_at_obs, 97.5, axis=0)
    cov_rep = coverage_from_ci(dy['reps_by_time'], dy['time'], lo, hi)
    cov_mean = float(np.mean((dy['mean'] >= lo) & (dy['mean'] <= hi)))
    ci_width_mean = float(np.mean(hi - lo))

    return {
        'dy_tmcmc_dir': dy_tmcmc['dir'],
        'sigma_knots_dir': meta.get('sigma_knots_dir', ''),
        'dynamic_c': True,
        'c_scale': c_scale,
        'dt': dt,
        'n_newton': n_newton,
        'time_to_step': time_to_step,
        'n_obs': int(len(dy['time'])),
        'k_MAP': k_map,
        'k_mean': float(np.mean(samples_k)),
        'k_std': float(np.std(samples_k)),
        'k_ci_lo': float(np.percentile(samples_k, 2.5)),
        'k_ci_hi': float(np.percentile(samples_k, 97.5)),
        'affine_a_map': float(coeff_map[0]),
        'affine_b_map': float(coeff_map[1]),
        **m,
        'pp_n_samples': int(len(k_sel)),
        'pp_coverage_reps_95': cov_rep,
        'pp_coverage_mean_95': cov_mean,
        'pp_ci_width_mean': ci_width_mean,
        'noise_floor': float(noise_floor),
    }


def plot_8panel_fast(out_png, tmcmc_dir, exp_data, dynamic_c, c_scale, dt, n_newton,
                     n_pp_samples=200, seed=0, noise_floor=0.1):
    rng = np.random.default_rng(seed)
    fig = plt.figure(figsize=(20, 8))
    gs = gridspec.GridSpec(2, 4, figure=fig, hspace=0.55, wspace=0.32)

    for idx_temp, temp in enumerate(TEMPS):
        ax = fig.add_subplot(gs[idx_temp // 4, idx_temp % 4])
        d = exp_data[temp]
        n_steps = MAX_STEPS[temp]
        npz = np.load(os.path.join(tmcmc_dir, f'samples_{temp}C.npz'))
        sigma_map = float(npz['sigma_MAP'])
        samples = npz['samples']

        phibar_map = simulate_phibar(sigma_map, n_steps, dynamic_c, c_scale, dt, n_newton)
        pred_map_obs, coeff = profiled_predictions(phibar_map, d['time'], d['mean'], n_steps)
        mapped_map_full = coeff[0] * phibar_map + coeff[1]
        t_full = np.arange(n_steps + 1) / TIME_TO_STEP
        ax.plot(t_full, mapped_map_full, '-', color='steelblue', lw=1.8, zorder=3, label='MAP fit')

        sel = rng.choice(len(samples), size=min(n_pp_samples, len(samples)), replace=False)
        lo, hi, _ = posterior_predictive_ci_obs(
            samples[sel], d['time'], d['mean'], n_steps, dynamic_c, c_scale, dt, n_newton
        )
        ax.fill_between(d['time'], lo, hi, alpha=0.25, color='steelblue', label='95% CI')

        for t_pt, reps in d['reps_by_time'].items():
            for v in reps:
                ax.plot(t_pt, v, 'o', color='#e05a2b', ms=4, alpha=0.85, zorder=4,
                        markeredgewidth=0.4, markeredgecolor='white')
        ax.errorbar(d['time'], d['mean'], yerr=d['std'],
                    fmt='none', ecolor='#e05a2b', elinewidth=1.0,
                    capsize=2, zorder=5, alpha=0.7)

        if dynamic_c:
            title_mode = f"c(t)={c_scale:g}*(1-φψ)"
        else:
            title_mode = "c=100"
        ax.set_title(f'({list("abcdefgh")[idx_temp]}) {temp}°C   MAP a₁₁={sigma_map:.2f}   {title_mode}',
                     fontsize=10, pad=5)
        ax.set_xlabel('Time (min)', fontsize=9)
        ax.set_ylabel('log₁₀(CFU/mL)', fontsize=9)
        ax.tick_params(labelsize=8)
        ax.spines[['top', 'right']].set_visible(False)

        if idx_temp == 0:
            import matplotlib.lines
            import matplotlib.patches
            ax.legend(fontsize=7, framealpha=0.7,
                      handles=[
                          matplotlib.lines.Line2D([], [], color='steelblue', lw=1.8, label='MAP fit'),
                          matplotlib.patches.Patch(color='steelblue', alpha=0.25, label='95% CI'),
                          matplotlib.lines.Line2D([], [], marker='o', color='w',
                                                  markerfacecolor='#e05a2b', ms=5, label='Experiment'),
                      ])

    fig.suptitle('Monospecies Hamilton ODE — TMCMC posterior fit', fontsize=12, y=1.01)
    fig.savefig(out_png, dpi=300, bbox_inches='tight')
    plt.close(fig)


def read_run_meta(tmcmc_dir):
    meta = {}
    json_path = os.path.join(tmcmc_dir, 'tmcmc_summary.json')
    if os.path.exists(json_path):
        with open(json_path, 'r') as f:
            meta = json.load(f)
    return meta


def process_one_dir(tmcmc_dir, exp_data, out_dir=None, n_pp_samples=200, seed=0, noise_floor=0.1,
                    override_dynamic_c=None, override_c_scale=None, override_dt=None, override_n_newton=None):
    tmcmc_dir = str(tmcmc_dir)
    meta = read_run_meta(tmcmc_dir)

    name_hint = os.path.basename(os.path.normpath(tmcmc_dir)).lower()
    if override_dynamic_c is None:
        if 'dynamic_c' in meta:
            dynamic_c = bool(meta['dynamic_c'])
        else:
            dynamic_c = ('dyncc' in name_hint) and ('staticc' not in name_hint)
    else:
        dynamic_c = bool(override_dynamic_c)

    if override_c_scale is None:
        if 'c_scale' in meta:
            c_scale = float(meta['c_scale'])
        else:
            if dynamic_c:
                import re
                m = re.search(r'dyncc([0-9]+(?:\.[0-9]+)?)', name_hint)
                c_scale = float(m.group(1)) if m else 10.0
            else:
                c_scale = 100.0
    else:
        c_scale = float(override_c_scale)

    dt = float(meta.get('dt', 1e-4)) if override_dt is None else float(override_dt)
    n_newton = int(meta.get('n_newton', 6)) if override_n_newton is None else int(override_n_newton)

    mutation = meta.get('mutation', None)
    n_particles = meta.get('n_particles', None)

    out_dir = tmcmc_dir if out_dir is None else str(out_dir)
    Path(out_dir).mkdir(parents=True, exist_ok=True)

    out_png = os.path.join(out_dir, f'monospecies_8panel_dyncc{c_scale:g}.png' if dynamic_c else 'monospecies_8panel_staticc100.png')
    plot_8panel_fast(out_png, tmcmc_dir, exp_data, dynamic_c, c_scale, dt, n_newton,
                     n_pp_samples=n_pp_samples, seed=seed, noise_floor=noise_floor)

    rows = []
    for temp in TEMPS:
        d = exp_data[temp]
        n_steps = MAX_STEPS[temp]
        npz = np.load(os.path.join(tmcmc_dir, f'samples_{temp}C.npz'))
        sigma_map = float(npz['sigma_MAP'])
        samples = npz['samples']

        phibar_map = simulate_phibar(sigma_map, n_steps, dynamic_c, c_scale, dt, n_newton)
        pred_map_obs, coeff = profiled_predictions(phibar_map, d['time'], d['mean'], n_steps)
        m = compute_metrics_at_obs(pred_map_obs, d['mean'], d['std'], noise_floor)

        rng = np.random.default_rng(seed + temp)
        sel = rng.choice(len(samples), size=min(n_pp_samples, len(samples)), replace=False)
        lo, hi, _ = posterior_predictive_ci_obs(
            samples[sel], d['time'], d['mean'], n_steps, dynamic_c, c_scale, dt, n_newton
        )
        cov_rep = coverage_from_ci(d['reps_by_time'], d['time'], lo, hi)
        cov_mean = float(np.mean((d['mean'] >= lo) & (d['mean'] <= hi)))
        ci_width_mean = float(np.mean(hi - lo))

        rows.append({
            'tmcmc_dir': tmcmc_dir,
            'mutation': mutation,
            'n_particles': n_particles,
            'dynamic_c': dynamic_c,
            'c_scale': c_scale,
            'dt': dt,
            'n_newton': n_newton,
            'temp_C': temp,
            'sigma_ref_felix': float(FELIX_SIGMA[temp]),
            'sigma_paper': float(PAPER_SIGMA[temp]),
            'sigma_MAP': sigma_map,
            'affine_a': float(coeff[0]),
            'affine_b': float(coeff[1]),
            **m,
            'pp_coverage_reps_95': cov_rep,
            'pp_coverage_mean_95': cov_mean,
            'pp_ci_width_mean': ci_width_mean,
            'n_obs': int(len(d['time'])),
            'n_pp_samples': int(min(n_pp_samples, len(samples))),
            'plot_8panel': out_png,
        })

    out_csv = os.path.join(out_dir, 'monospecies_compiled_metrics.csv')
    fieldnames = list(rows[0].keys()) if rows else []
    with open(out_csv, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)

    return out_csv, out_png, rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--tmcmc_dir', action='append', required=True,
                        help='TMCMC output directory (repeatable)')
    parser.add_argument('--out_csv', default=None,
                        help='Master CSV path. Default: <monospecies_root>/monospecies_master_compiled_metrics.csv')
    parser.add_argument('--n_pp_samples', type=int, default=200)
    parser.add_argument('--seed', type=int, default=0)
    parser.add_argument('--noise_floor', type=float, default=0.1)
    parser.add_argument('--xlsx', default=os.path.join(os.path.dirname(__file__), 'raw data.xlsx'))
    parser.add_argument('--analyze_dynamic', action='store_true')
    parser.add_argument('--dy_time_to_step', type=float, default=10.0)
    parser.add_argument('--dy_sheet', default='Dy')
    parser.add_argument('--dy_tmcmc_dir', type=str, default=None,
                        help='Directory containing tmcmc_summary_Dy_k.csv/json and samples_Dy_k.npz')
    parser.add_argument('--dy_sigma_knots_dir', type=str, default=None,
                        help='Directory containing samples_{T}C.npz used as sigma knots for Dy plotting')
    parser.add_argument('--dy_n_pp_samples', type=int, default=200)
    args = parser.parse_args()

    if args.dy_tmcmc_dir is not None:
        dy = load_dynamic_data_with_reps(args.xlsx, sheet=args.dy_sheet)
        dy_tmcmc = load_dy_tmcmc_outputs(args.dy_tmcmc_dir)

        sigma_knots_dir = args.dy_sigma_knots_dir
        if sigma_knots_dir is None:
            sigma_knots_dir = dy_tmcmc['meta'].get('sigma_knots_dir', None)
        if sigma_knots_dir is None:
            parent_tmcmc = os.path.join(os.path.dirname(args.dy_tmcmc_dir), '_tmcmc_results')
            sigma_knots_dir = parent_tmcmc if os.path.isdir(parent_tmcmc) else None

        if sigma_knots_dir is not None and os.path.isdir(str(sigma_knots_dir)):
            sigma_knots = load_sigma_knots_from_dir(sigma_knots_dir)
        else:
            sigma_knots = {float(t): float(FELIX_SIGMA[int(t)]) for t in TEMPS}

        out_base = os.path.join(args.dy_tmcmc_dir, 'liu2019_fig1_dynamic_tmcmc')
        plot_dy_tmcmc_fig1_like(
            out_base,
            dy=dy,
            dy_tmcmc=dy_tmcmc,
            sigma_knots=sigma_knots,
            n_pp_samples=args.dy_n_pp_samples,
            seed=args.seed,
        )
        plot_dy_k_posterior(os.path.join(args.dy_tmcmc_dir, 'dy_k_posterior'), dy_tmcmc)
        metrics = evaluate_dy_tmcmc_metrics(
            dy,
            dy_tmcmc=dy_tmcmc,
            sigma_knots=sigma_knots,
            n_pp_samples=max(args.dy_n_pp_samples, 200),
            seed=args.seed,
            noise_floor=args.noise_floor,
        )
        out_csv = os.path.join(args.dy_tmcmc_dir, 'dy_metrics.csv')
        with open(out_csv, 'w', newline='') as f:
            w = csv.DictWriter(f, fieldnames=list(metrics.keys()))
            w.writeheader()
            w.writerow(metrics)
        out_json = os.path.join(args.dy_tmcmc_dir, 'dy_metrics.json')
        with open(out_json, 'w') as f:
            json.dump(metrics, f, indent=2)
        print(f"Saved: {out_csv}")
        print(f"Saved: {out_json}")
        print(f"Saved: {out_base}.pdf/.png")
        print(f"Saved: {os.path.join(args.dy_tmcmc_dir, 'dy_k_posterior.pdf/.png')}")
        return

    exp_data = load_static_data_with_reps(args.xlsx)
    dy = load_dynamic_data_with_reps(args.xlsx, sheet=args.dy_sheet) if args.analyze_dynamic else None

    all_rows = []
    for d in args.tmcmc_dir:
        _, _, rows = process_one_dir(
            d, exp_data,
            out_dir=d,
            n_pp_samples=args.n_pp_samples,
            seed=args.seed,
            noise_floor=args.noise_floor,
        )
        all_rows.extend(rows)

        if dy is not None:
            sigma_knots = {}
            for temp in TEMPS:
                npz = np.load(os.path.join(d, f'samples_{temp}C.npz'))
                sigma_knots[float(temp)] = float(npz['sigma_MAP'])
            meta = read_run_meta(d)
            dynamic_c = bool(meta.get('dynamic_c', True))
            c_scale = float(meta.get('c_scale', 10.0))
            dt = float(meta.get('dt', 1e-4))
            n_newton = int(meta.get('n_newton', 6))
            fit = fit_dynamic_scale_k(
                dy,
                sigma_knots=sigma_knots,
                dt=dt,
                n_newton=n_newton,
                dynamic_c=dynamic_c,
                c_scale=c_scale,
                time_to_step_dy=args.dy_time_to_step,
                noise_floor=args.noise_floor,
            )
            out_png = os.path.join(d, 'liu2019_fig1_dynamic_fit.png')
            plot_dynamic_overview(out_png, dy, fit=fit)
            print(f"Saved: {out_png}")

    out_csv = args.out_csv
    if out_csv is None:
        out_csv = os.path.join(os.path.dirname(__file__), 'monospecies_master_compiled_metrics.csv')
    fieldnames = list(all_rows[0].keys()) if all_rows else []
    with open(out_csv, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(all_rows)
    print(f"Saved: {out_csv}")


if __name__ == '__main__':
    main()
