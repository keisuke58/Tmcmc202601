#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_synthetic_rawdata.py
─────────────────────────────
TMCMC MAP a₁₁ 推定値から "raw data.xlsx" 相当の合成データを生成する。

処理内容:
  1. dyncc10 TMCMC 結果から MAP a₁₁ を読み込む
  2. Hamilton ODE (c(t)=c_scale*(1-φψ)) を各温度で積分
  3. 元データの timepoints & scale に合わせた affine mapping を fitting
  4. 元データの実験ノイズ (std) を推定して合成 3 replicate を生成
  5. 元 xlsx と同フォーマット (St sheet) で出力

Usage:
    python generate_synthetic_rawdata.py \
        --tmcmc_dir _tmcmc_results_dyncc10_rw4000p \
        --outfile synthetic_raw_data.xlsx \
        [--noise_seed 42] [--noise_scale 1.0]
"""

import os
import sys
import argparse
import numpy as np

sys.path.insert(0, os.path.dirname(__file__))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'main'))

from hamilton_monospecies import simulate_felix_dynamic_c_exogenous_scale

TEMPS   = [4, 8, 15, 20, 25, 35, 37, 40]
MAX_STEPS = {4: 4000, 8: 2000, 15: 1000, 20: 500,
             25: 300,  35: 200,  37: 200, 40: 450}
TIME_TO_STEP = 10.0


# ── データ読み込み ──────────────────────────────────────────────────────────

def load_original_data(xlsx_path):
    """元 xlsx の St シートから全レプリケート + 時系列を読む。"""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['St']

    data = {}
    for i, temp in enumerate(TEMPS):
        col_start = i * 3  # 0-indexed column offsets
        times_rep, vals_rep = [], []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            t = row[col_start]
            v = row[col_start + 1]
            if t is not None and v is not None:
                times_rep.append(float(t))
                vals_rep.append(float(v))

        # グループ化（同じ timepoint → 3 replicate）
        from collections import defaultdict
        reps = defaultdict(list)
        for t, v in zip(times_rep, vals_rep):
            reps[t].append(v)

        unique_times = sorted(reps.keys())
        data[temp] = {
            'times': unique_times,
            'reps': {t: reps[t] for t in unique_times},
            'mean': np.array([np.mean(reps[t]) for t in unique_times]),
            'std':  np.array([np.std(reps[t])  for t in unique_times]),
            'n_reps': max(len(reps[t]) for t in unique_times),
        }
    return data


# ── ODE 積分 ────────────────────────────────────────────────────────────────

def simulate_trajectory(a11, n_steps, c_scale=10.0, dt=1e-4, n_newton=6, phi_init=0.1):
    """phibar(t) = phi(t)*psi(t) を返す (長さ n_steps+1)。"""
    _, _, phibar = simulate_felix_dynamic_c_exogenous_scale(
        a11, n_steps,
        dt=dt, Kp=1e-4, eta=1.0, eta_phi=1.0,
        c_scale=float(c_scale), alpha=0.0, n_newton=n_newton,
        phi_init=phi_init,
    )
    return np.asarray(phibar)


# ── Affine mapping fitting ───────────────────────────────────────────────────

def fit_affine(phibar, obs_times, obs_mean, n_steps):
    """
    MAP trajectory → log10(CFU) の affine map を最小二乗フィット。
    Returns (coeff [a, b], pred_at_obs)
    """
    step_idx = np.clip(
        (np.array(obs_times) * TIME_TO_STEP).astype(int), 0, n_steps
    )
    sim_vals = phibar[step_idx]
    X = np.column_stack([sim_vals, np.ones_like(sim_vals)])
    coeff, _, _, _ = np.linalg.lstsq(X, obs_mean, rcond=None)
    pred = X @ coeff
    return coeff, pred


# ── 合成レプリケート生成 ─────────────────────────────────────────────────────

def make_synthetic_replicates(pred_at_obs, orig_std, n_reps, rng, noise_scale=1.0):
    """
    各観測時刻で pred_at_obs ± 元データ std に基づく合成レプリケートを生成。
    noise_scale=1.0 → 元と同等のばらつき。
    Returns list of length n_obs, each being list of n_reps values.
    """
    result = []
    for mu, sigma in zip(pred_at_obs, orig_std):
        # std が 0 の場合は最小値 0.05 を使う
        s = max(float(sigma) * noise_scale, 0.05)
        reps = rng.normal(loc=mu, scale=s, size=n_reps).tolist()
        result.append(reps)
    return result


# ── Excel 出力 ───────────────────────────────────────────────────────────────

def write_xlsx(out_path, synth_data):
    """
    synth_data: dict {temp → {'times': [...], 'reps': [[r1,r2,r3], ...], 'mean': ..., 'std': ...}}
    元 xlsx と同フォーマット:
      Row1: 温度ヘッダー
      Row2: Time / Con カラムヘッダー
      Row3+: 各レプリケート行（同 timepoint が n_reps 行）
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'St'

    # ─── Row 1: 温度ヘッダー ───
    for i, temp in enumerate(TEMPS):
        col = i * 3 + 1  # 1-indexed
        ws.cell(row=1, column=col, value=f'{temp}°C')
        ws.cell(row=1, column=col).font = Font(bold=True)

    # ─── Row 2: Time/Con ヘッダー ───
    for i, temp in enumerate(TEMPS):
        col = i * 3 + 1
        ws.cell(row=2, column=col,   value='Time')
        ws.cell(row=2, column=col+1, value='Con')
        for c in [col, col+1]:
            ws.cell(row=2, column=c).font = Font(bold=True)

    # ─── Row 3+: データ ───
    # 全温度の最大行数を計算
    max_rows = max(
        len(synth_data[t]['times']) * len(synth_data[t]['reps'][synth_data[t]['times'][0]])
        for t in TEMPS
    )

    # 各温度のフラット行リスト [(time, con), ...]
    flat = {}
    for temp in TEMPS:
        td = synth_data[temp]
        rows = []
        for tp in td['times']:
            for v in td['reps'][tp]:
                rows.append((tp, round(v, 2)))
        flat[temp] = rows

    max_data_rows = max(len(flat[t]) for t in TEMPS)
    for r in range(max_data_rows):
        for i, temp in enumerate(TEMPS):
            col = i * 3 + 1
            fd = flat[temp]
            if r < len(fd):
                t_val, c_val = fd[r]
                ws.cell(row=3 + r, column=col,   value=t_val)
                ws.cell(row=3 + r, column=col+1, value=c_val)

    # 列幅調整
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 10

    wb.save(out_path)
    print(f'Saved: {out_path}')


# ── メイン ───────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--tmcmc_dir', type=str,
                        default='_tmcmc_results_dyncc10_rw4000p',
                        help='TMCMC 結果ディレクトリ (samples_TC.npz を含む)')
    parser.add_argument('--orig_xlsx', type=str,
                        default='raw data.xlsx',
                        help='元 xlsx (timepoints とノイズ推定に使用)')
    parser.add_argument('--outfile', type=str,
                        default='synthetic_raw_data.xlsx')
    parser.add_argument('--c_scale', type=float, default=10.0)
    parser.add_argument('--dt', type=float, default=1e-4)
    parser.add_argument('--n_newton', type=int, default=6)
    parser.add_argument('--phi_init', type=float, default=0.1)
    parser.add_argument('--noise_scale', type=float, default=1.0,
                        help='元データの std に対するスケール (1.0 = 同等)')
    parser.add_argument('--noise_seed', type=int, default=42)
    parser.add_argument('--no_noise', action='store_true',
                        help='ノイズなし (3 rep すべて同一 MAP 値)')
    args = parser.parse_args()

    rng = np.random.default_rng(args.noise_seed)

    # 1. 元データ読み込み
    xlsx_path = os.path.join(os.path.dirname(__file__), args.orig_xlsx)
    print(f'Loading original data: {xlsx_path}')
    orig = load_original_data(xlsx_path)

    # 2. MAP a₁₁ 読み込み
    print(f'Loading MAP a₁₁ from: {args.tmcmc_dir}')
    a11_map = {}
    for temp in TEMPS:
        npz_path = os.path.join(
            os.path.dirname(__file__), args.tmcmc_dir, f'samples_{temp}C.npz'
        )
        if not os.path.exists(npz_path):
            raise FileNotFoundError(f'Missing: {npz_path}')
        d = np.load(npz_path)
        a11_map[temp] = float(d['sigma_MAP'])
        print(f'  {temp:>3d}°C  a₁₁ = {a11_map[temp]:.4f}')

    # 3. 各温度で ODE → affine fitting → 合成レプリケート
    synth_data = {}
    print('\n--- Fitting affine map & generating replicates ---')
    print(f"{'Temp':>5s}  {'a11':>8s}  {'a(slope)':>10s}  {'b(offset)':>10s}  "
          f"{'RMSE':>8s}  {'n_obs':>6s}")
    print('-' * 60)

    for temp in TEMPS:
        od = orig[temp]
        n_steps = MAX_STEPS[temp]
        a11 = a11_map[temp]

        # ODE 積分
        phibar = simulate_trajectory(
            a11, n_steps, c_scale=args.c_scale,
            dt=args.dt, n_newton=args.n_newton, phi_init=args.phi_init,
        )

        # Affine fitting
        coeff, pred_at_obs = fit_affine(phibar, od['times'], od['mean'], n_steps)
        a_slope, b_offset = float(coeff[0]), float(coeff[1])
        rmse = float(np.sqrt(np.mean((pred_at_obs - od['mean'])**2)))

        # 合成レプリケート
        n_reps = od['n_reps']
        if args.no_noise:
            reps_list = [[float(round(v, 3))] * n_reps for v in pred_at_obs]
        else:
            reps_list = make_synthetic_replicates(
                pred_at_obs, od['std'], n_reps, rng, noise_scale=args.noise_scale
            )

        synth_data[temp] = {
            'times': od['times'],
            'reps':  {t: reps for t, reps in zip(od['times'], reps_list)},
            'mean':  np.array([np.mean(r) for r in reps_list]),
            'std':   np.array([np.std(r)  for r in reps_list]),
        }

        print(f"  {temp:>3d}°C  {a11:>8.3f}  {a_slope:>10.4f}  {b_offset:>10.4f}  "
              f"{rmse:>8.4f}  {len(od['times']):>6d}")

    # 4. Excel 出力
    out_path = os.path.join(os.path.dirname(__file__), args.outfile)
    write_xlsx(out_path, synth_data)

    # 5. 比較サマリ
    print('\n--- Synthetic vs Original (mean log10 CFU/mL) ---')
    for temp in TEMPS:
        od = orig[temp]
        sd = synth_data[temp]
        print(f'\n  {temp}°C:')
        print(f"    {'Time':>6s}  {'Orig mean':>10s}  {'Synth mean':>11s}  {'Diff':>8s}")
        for t, om, sm in zip(od['times'], od['mean'], sd['mean']):
            print(f"    {t:>6.0f}  {om:>10.3f}  {sm:>11.3f}  {sm-om:>+8.4f}")


if __name__ == '__main__':
    main()
