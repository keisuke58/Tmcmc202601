# -*- coding: utf-8 -*-
"""
plot_monospecies_8panel.py — 8-panel figure for monospecies TMCMC results.

各温度 (4,8,15,20,25,35,37,40°C) について:
  - 実験データ (3 replicates) の散布点
  - MAP σ による Hamilton モデルフィット (affine mapping)
  - TMCMC サンプルによる posterior predictive 95% CI (shaded)

Usage:
    python plot_monospecies_8panel.py
    python plot_monospecies_8panel.py --outfile monospecies_8panel.pdf
"""

import os
import sys
import argparse
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec

sys.path.insert(0, os.path.dirname(__file__))

import jax
jax.config.update('jax_enable_x64', True)
import jax.numpy as jnp

from hamilton_monospecies_jax import (
    simulate_monospecies_phibar_hist,
    simulate_monospecies_phibar_hist_dynamic_c,
)

# ── Constants ────────────────────────────────────────────────────────────────

TEMPS = [4, 8, 15, 20, 25, 35, 37, 40]
MAX_STEPS = {4: 4000, 8: 2000, 15: 1000, 20: 500,
             25: 300,  35: 200,  37: 200, 40: 450}
TIME_TO_STEP = 10.0   # 1 min → 10 steps
TMCMC_DIR = os.path.join(os.path.dirname(__file__), '_tmcmc_results')
DATA_XLSX = os.path.join(os.path.dirname(__file__), 'raw data.xlsx')

PANEL_LABELS = list('abcdefgh')


# ── Data loading ─────────────────────────────────────────────────────────────

def load_static_data(xlsx_path):
    """Load St sheet: all replicates per temperature."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['St']
    data = {}
    for i, temp in enumerate(TEMPS):
        col_start = i * 3
        times_rep, vals_rep = [], []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            t_val = row[col_start]
            c_val = row[col_start + 1]
            if t_val is not None and c_val is not None:
                times_rep.append(float(t_val))
                vals_rep.append(float(c_val))
        # group into 3 replicates per timepoint
        unique_times = sorted(set(times_rep))
        reps_by_time = {t: [] for t in unique_times}
        for t, v in zip(times_rep, vals_rep):
            reps_by_time[t].append(v)
        t_arr = np.array(unique_times)
        mean_arr = np.array([np.mean(reps_by_time[t]) for t in unique_times])
        std_arr  = np.array([np.std(reps_by_time[t])  for t in unique_times])
        data[temp] = {
            'time': t_arr,
            'mean': mean_arr,
            'std':  std_arr,
            'reps_by_time': reps_by_time,
        }
    return data


# ── Simulation helper ─────────────────────────────────────────────────────────

def affine_fit(sim_vals, obs_vals):
    """Least-squares affine map: obs ≈ a*sim + b."""
    X = np.column_stack([sim_vals, np.ones_like(sim_vals)])
    coeff, _, _, _ = np.linalg.lstsq(X, obs_vals, rcond=None)
    return coeff  # [a, b]


def simulate_and_map(sigma, n_steps, obs_time, obs_mean,
                     Kp=1e-4, eta=1.0, eta_phi=1.0, c=100.0,
                     dynamic_c=False, c_scale=10.0, dt=1e-4, n_newton=6):
    """Run ODE, extract at obs_time, fit affine map, return full mapped curve."""
    if dynamic_c:
        phibar = np.array(simulate_monospecies_phibar_hist_dynamic_c(
            float(sigma), n_steps, dt=dt, Kp=Kp, eta=eta, eta_phi=eta_phi, c_scale=c_scale, n_newton=n_newton
        ))
    else:
        phibar = np.array(simulate_monospecies_phibar_hist(
            float(sigma), n_steps, dt=dt, Kp=Kp, eta=eta, eta_phi=eta_phi, c=c, n_newton=n_newton
        ))
    # extract at obs timepoints
    idx = np.clip((obs_time * TIME_TO_STEP).astype(int), 0, n_steps)
    sim_at_obs = phibar[idx]
    coeff = affine_fit(sim_at_obs, obs_mean)
    # mapped full curve
    t_full = np.arange(n_steps + 1) / TIME_TO_STEP  # minutes
    mapped_full = coeff[0] * phibar + coeff[1]
    return t_full, mapped_full, coeff


# ── Main plotting ─────────────────────────────────────────────────────────────

def plot_8panel(outfile='monospecies_8panel.png', n_pp_samples=200, seed=0,
                tmcmc_dir=TMCMC_DIR, dynamic_c=False, c_scale=10.0, dt=1e-4, n_newton=6):
    rng = np.random.default_rng(seed)

    exp_data = load_static_data(DATA_XLSX)

    fig = plt.figure(figsize=(14, 10))
    gs = gridspec.GridSpec(2, 4, figure=fig, hspace=0.45, wspace=0.35)

    for idx, temp in enumerate(TEMPS):
        ax = fig.add_subplot(gs[idx // 4, idx % 4])
        d = exp_data[temp]
        n_steps = MAX_STEPS[temp]

        # ── Load TMCMC results ──────────────────────────────────────────────
        npz_path = os.path.join(tmcmc_dir, f'samples_{temp}C.npz')
        if os.path.exists(npz_path):
            npz = np.load(npz_path)
            sigma_MAP = float(npz['sigma_MAP'])
            samples   = npz['samples']
        else:
            # fallback: use Felix reference sigma
            from estimate_monospecies_tmcmc import FELIX_SIGMA
            sigma_MAP = FELIX_SIGMA[temp]
            samples   = np.array([sigma_MAP])

        # ── Posterior predictive CI ─────────────────────────────────────────
        sel = rng.choice(len(samples), size=min(n_pp_samples, len(samples)),
                         replace=False)
        pp_curves = []
        for s_val in samples[sel]:
            try:
                t_full, mapped, _ = simulate_and_map(
                    s_val, n_steps, d['time'], d['mean'],
                    dynamic_c=dynamic_c, c_scale=c_scale, dt=dt, n_newton=n_newton)
                pp_curves.append(mapped)
            except Exception:
                pass

        if pp_curves:
            pp_stack = np.stack(pp_curves, axis=0)
            pp_lo = np.percentile(pp_stack, 2.5,  axis=0)
            pp_hi = np.percentile(pp_stack, 97.5, axis=0)
            # t_full same for all
            t_full, _, _ = simulate_and_map(
                sigma_MAP, n_steps, d['time'], d['mean'],
                dynamic_c=dynamic_c, c_scale=c_scale, dt=dt, n_newton=n_newton)
            ax.fill_between(t_full, pp_lo, pp_hi,
                            alpha=0.25, color='steelblue', label='95% CI')

        # ── MAP fit ─────────────────────────────────────────────────────────
        try:
            t_full, mapped_map, _ = simulate_and_map(
                sigma_MAP, n_steps, d['time'], d['mean'],
                dynamic_c=dynamic_c, c_scale=c_scale, dt=dt, n_newton=n_newton)
            ax.plot(t_full, mapped_map, '-', color='steelblue',
                    lw=1.8, zorder=3, label=f'MAP fit')
        except Exception as e:
            print(f"  {temp}°C MAP failed: {e}")

        # ── Experimental data ────────────────────────────────────────────────
        for t_pt, reps in d['reps_by_time'].items():
            for v in reps:
                ax.plot(t_pt, v, 'o', color='#e05a2b', ms=4,
                        alpha=0.85, zorder=4, markeredgewidth=0.4,
                        markeredgecolor='white')

        # mean ± std error bar
        ax.errorbar(d['time'], d['mean'], yerr=d['std'],
                    fmt='none', ecolor='#e05a2b', elinewidth=1.0,
                    capsize=2, zorder=5, alpha=0.7)

        # ── Formatting ───────────────────────────────────────────────────────
        if dynamic_c:
            title_mode = f"c(t)={c_scale:g}*(1-φψ)"
        else:
            title_mode = "c=100"
        ax.set_title(f'({PANEL_LABELS[idx]}) {temp}°C   σ={sigma_MAP:.2f}   {title_mode}',
                     fontsize=9, pad=4)
        ax.set_xlabel('Time (min)', fontsize=8)
        ax.set_ylabel('log₁₀(CFU/mL)', fontsize=8)
        ax.tick_params(labelsize=7)
        ax.spines[['top', 'right']].set_visible(False)

        # legend only on first panel
        if idx == 0:
            ax.legend(fontsize=7, framealpha=0.7,
                      handles=[
                          matplotlib.lines.Line2D([], [], color='steelblue',
                                                  lw=1.8, label='MAP fit'),
                          matplotlib.patches.Patch(color='steelblue', alpha=0.25,
                                                   label='95% CI'),
                          matplotlib.lines.Line2D([], [], marker='o', color='w',
                                                  markerfacecolor='#e05a2b',
                                                  ms=5, label='Experiment'),
                      ])

    fig.suptitle('Monospecies Hamilton ODE — TMCMC posterior fit\n'
                 '(8 temperature conditions)',
                 fontsize=11, y=1.01)
    fig.savefig(outfile, dpi=150, bbox_inches='tight')
    print(f"Saved: {outfile}")


def plot_4c8c_md_reference(outfile='monospecies_4C8C_md_dynamic_c.png', c_scale=10.0, dt=1e-4, n_newton=6):
    from estimate_monospecies_tmcmc import FELIX_SIGMA

    exp_data = load_static_data(DATA_XLSX)
    fig, axes = plt.subplots(1, 2, figsize=(10.5, 4.0), constrained_layout=True)

    for ax, temp, panel in zip(axes, [4, 8], ['(a)', '(b)']):
        d = exp_data[temp]
        n_steps = MAX_STEPS[temp]
        sigma = float(FELIX_SIGMA[temp])

        t_full, mapped, _ = simulate_and_map(
            sigma, n_steps, d['time'], d['mean'], dynamic_c=True, c_scale=c_scale, dt=dt, n_newton=n_newton
        )
        ax.plot(t_full, mapped, '-', color='steelblue', lw=2.0, zorder=3, label='MD fit')

        for t_pt, reps in d['reps_by_time'].items():
            for v in reps:
                ax.plot(t_pt, v, 'o', color='#e05a2b', ms=4, alpha=0.85, zorder=4,
                        markeredgewidth=0.4, markeredgecolor='white')

        ax.errorbar(d['time'], d['mean'], yerr=d['std'],
                    fmt='none', ecolor='#e05a2b', elinewidth=1.0,
                    capsize=2, zorder=5, alpha=0.7)

        ax.set_title(f'{panel} {temp}°C   σ={sigma:.2f}   c(t)={c_scale:g}*(1-φψ)', fontsize=9, pad=4)
        ax.set_xlabel('Time (min)', fontsize=8)
        ax.set_ylabel('log₁₀(CFU/mL)', fontsize=8)
        ax.tick_params(labelsize=7)
        ax.spines[['top', 'right']].set_visible(False)

    axes[0].legend(fontsize=7, framealpha=0.7,
                   handles=[
                       matplotlib.lines.Line2D([], [], color='steelblue', lw=2.0, label='MD fit'),
                       matplotlib.lines.Line2D([], [], marker='o', color='w',
                                               markerfacecolor='#e05a2b', ms=5, label='Experiment'),
                   ])

    fig.savefig(outfile, dpi=150, bbox_inches='tight')
    print(f"Saved: {outfile}")


def plot_4c8c_reestimate_tmcmc(tmcmc_dir, outfile='monospecies_4C8C_reest_tmcmc.png',
                              n_pp_samples=200, seed=0, dynamic_c=True, c_scale=100.0, dt=1e-4, n_newton=6):
    rng = np.random.default_rng(seed)
    exp_data = load_static_data(DATA_XLSX)
    fig, axes = plt.subplots(1, 2, figsize=(10.5, 4.0), constrained_layout=True)

    for ax, temp, panel in zip(axes, [4, 8], ['(a)', '(b)']):
        d = exp_data[temp]
        n_steps = MAX_STEPS[temp]
        npz_path = os.path.join(tmcmc_dir, f'samples_{temp}C.npz')
        npz = np.load(npz_path)
        sigma_MAP = float(npz['sigma_MAP'])
        samples = npz['samples']

        sel = rng.choice(len(samples), size=min(n_pp_samples, len(samples)), replace=False)
        pp_curves = None
        if len(sel) >= 2:
            sigmas = jnp.array(samples[sel], dtype=jnp.float64)
            obs_time = jnp.array(d['time'], dtype=jnp.float64)
            obs_mean = jnp.array(d['mean'], dtype=jnp.float64)
            idx = jnp.clip((obs_time * TIME_TO_STEP).astype(jnp.int32), 0, n_steps)

            def _simulate_one(s):
                if dynamic_c:
                    return simulate_monospecies_phibar_hist_dynamic_c(
                        s, n_steps, dt=dt, Kp=1e-4, eta=1.0, eta_phi=1.0, c_scale=c_scale, n_newton=n_newton
                    )
                return simulate_monospecies_phibar_hist(
                    s, n_steps, dt=dt, Kp=1e-4, eta=1.0, eta_phi=1.0, c=100.0, n_newton=n_newton
                )

            phibars = jax.vmap(_simulate_one)(sigmas)  # [B, T]
            x = phibars[:, idx]  # [B, n_obs]

            mx = jnp.mean(x, axis=1, keepdims=True)
            my = jnp.mean(obs_mean, axis=0, keepdims=True)
            xc = x - mx
            yc = obs_mean[None, :] - my
            var_x = jnp.mean(xc * xc, axis=1)
            cov_xy = jnp.mean(xc * yc, axis=1)
            denom = jnp.maximum(var_x, 1e-12)
            a = cov_xy / denom
            b = jnp.squeeze(my, axis=0) - a * jnp.squeeze(mx, axis=1)

            mapped = a[:, None] * phibars + b[:, None]
            pp_lo = jnp.percentile(mapped, 2.5, axis=0)
            pp_hi = jnp.percentile(mapped, 97.5, axis=0)
            t_full = np.arange(n_steps + 1) / TIME_TO_STEP
            ax.fill_between(t_full, np.array(pp_lo), np.array(pp_hi),
                            alpha=0.25, color='steelblue', label='95% CI')

        t_full, mapped_map, _ = simulate_and_map(
            sigma_MAP, n_steps, d['time'], d['mean'],
            dynamic_c=dynamic_c, c_scale=c_scale, dt=dt, n_newton=n_newton,
        )
        ax.plot(t_full, mapped_map, '-', color='steelblue', lw=2.0, zorder=3, label='MAP fit')

        for t_pt, reps in d['reps_by_time'].items():
            for v in reps:
                ax.plot(t_pt, v, 'o', color='#e05a2b', ms=4, alpha=0.85, zorder=4,
                        markeredgewidth=0.4, markeredgecolor='white')

        ax.errorbar(d['time'], d['mean'], yerr=d['std'],
                    fmt='none', ecolor='#e05a2b', elinewidth=1.0,
                    capsize=2, zorder=5, alpha=0.7)

        c_str = f"c(t)={c_scale:g}*(1-φψ)" if dynamic_c else "c=100"
        ax.set_title(f'{panel} {temp}°C   MAP σ={sigma_MAP:.2f}   {c_str}', fontsize=9, pad=4)
        ax.set_xlabel('Time (min)', fontsize=8)
        ax.set_ylabel('log₁₀(CFU/mL)', fontsize=8)
        ax.tick_params(labelsize=7)
        ax.spines[['top', 'right']].set_visible(False)

    axes[0].legend(fontsize=7, framealpha=0.7,
                   handles=[
                       matplotlib.lines.Line2D([], [], color='steelblue', lw=2.0, label='MAP fit'),
                       matplotlib.patches.Patch(color='steelblue', alpha=0.25, label='95% CI'),
                       matplotlib.lines.Line2D([], [], marker='o', color='w',
                                               markerfacecolor='#e05a2b', ms=5, label='Experiment'),
                   ])
    fig.savefig(outfile, dpi=150, bbox_inches='tight')
    print(f"Saved: {outfile}")


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import matplotlib.lines
    import matplotlib.patches

    parser = argparse.ArgumentParser()
    parser.add_argument('--outfile', default='monospecies_8panel.png')
    parser.add_argument('--outfile_md', default='monospecies_4C8C_md_dynamic_c.png')
    parser.add_argument('--outfile_reest', default='monospecies_4C8C_reest_dyncc100.png')
    parser.add_argument('--tmcmc_dir_reest', default='_tmcmc_results_dyncc100_reest_4C8C')
    parser.add_argument('--tmcmc_dir', default=TMCMC_DIR)
    parser.add_argument('--n_pp_samples', type=int, default=200,
                        help='Number of posterior samples for CI band')
    parser.add_argument('--seed', type=int, default=0)
    parser.add_argument('--c_scale', type=float, default=10.0)
    parser.add_argument('--dynamic_c', action='store_true')
    parser.add_argument('--dt', type=float, default=1e-4)
    parser.add_argument('--n_newton', type=int, default=6)
    parser.add_argument('--only_md', action='store_true')
    parser.add_argument('--plot_reest_4c8c', action='store_true')
    args = parser.parse_args()

    if not args.only_md:
        plot_8panel(outfile=args.outfile,
                    n_pp_samples=args.n_pp_samples,
                    seed=args.seed,
                    tmcmc_dir=args.tmcmc_dir,
                    dynamic_c=args.dynamic_c,
                    c_scale=args.c_scale,
                    dt=args.dt,
                    n_newton=args.n_newton)
    plot_4c8c_md_reference(outfile=args.outfile_md, c_scale=args.c_scale, dt=args.dt, n_newton=args.n_newton)
    if args.plot_reest_4c8c:
        plot_4c8c_reestimate_tmcmc(
            tmcmc_dir=args.tmcmc_dir_reest,
            outfile=args.outfile_reest,
            n_pp_samples=args.n_pp_samples,
            seed=args.seed,
            dynamic_c=True,
            c_scale=100.0,
            dt=args.dt,
            n_newton=args.n_newton,
        )
