# -*- coding: utf-8 -*-
"""
estimate_monospecies_tmcmc.py — TMCMC estimation of a₁₁ (σ) for monospecies biofilm.

Uses JAX Hamilton ODE + tmcmc_nuts_engine for 8 temperature conditions.
1 free parameter per condition: σ ∈ [1, 500].
Affine mapping φ̄(t) → log₁₀(CFU/mL) is profiled out analytically.

Usage:
    python estimate_monospecies_tmcmc.py [--mutation rw|nuts] [--n_particles 200]
    python estimate_monospecies_tmcmc.py --temp 37   # single condition
"""

import os
import sys
import argparse
import time
import json
import csv
import numpy as np

def _parse_device_early():
    for i, a in enumerate(sys.argv):
        if a == "--device" and i + 1 < len(sys.argv):
            return sys.argv[i + 1]
    return "auto"


_device_early = _parse_device_early()
if _device_early == "cpu":
    os.environ["CUDA_VISIBLE_DEVICES"] = ""
    os.environ.setdefault("JAX_PLATFORMS", "cpu")
elif _device_early == "gpu":
    os.environ.setdefault("JAX_PLATFORMS", "cuda")
    try:
        import jax_cuda12_plugin  # noqa: F401
    except ImportError:
        pass


import jax
import jax.numpy as jnp

jax.config.update('jax_enable_x64', True)

# Add parent paths
sys.path.insert(0, os.path.dirname(__file__))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'main'))

from hamilton_monospecies_jax import (
    simulate_monospecies_phibar_hist,
    simulate_monospecies_phibar_hist_dynamic_c,
    simulate_monospecies_phibar_hist_dynamic_c_sigma_series,
)

# Felix's reference values (felix_monospecies_dynamic_c_params.md)
# Used as prior center for each temperature condition.
FELIX_SIGMA = {4: 4.25, 8: 10.0, 15: 25.0, 20: 50.0,
               25: 100.0, 35: 110.0, 37: 115.0, 40: 40.0}


def load_static_data(xlsx_path):
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['St']
    temps = [4, 8, 15, 20, 25, 35, 37, 40]
    data = {}
    for i, temp in enumerate(temps):
        col_start = i * 3
        times_all, cfu_all = [], []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
            t_cell, c_cell = row[col_start], row[col_start + 1]
            if t_cell.value is not None and c_cell.value is not None:
                times_all.append(float(t_cell.value))
                cfu_all.append(float(c_cell.value))
        unique_times = sorted(set(times_all))
        avg_cfu = [np.mean([cfu_all[j] for j in range(len(times_all)) if times_all[j] == t])
                   for t in unique_times]
        data[temp] = {'time': np.array(unique_times), 'cfu_mean': np.array(avg_cfu)}
    return data


def load_dynamic_data_with_reps(xlsx_path, sheet='Dy'):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet]

    obs_time, obs_vals = [], []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        t, r1, r2, r3 = row[0], row[1], row[2], row[3]
        if t is None:
            continue
        reps = [v for v in (r1, r2, r3) if v is not None]
        if not reps:
            continue
        obs_time.append(float(t))
        obs_vals.append([float(v) for v in reps])

    unique_times = sorted(set(obs_time))
    reps_by_time = {t: [] for t in unique_times}
    for t, reps in zip(obs_time, obs_vals):
        reps_by_time[float(t)].extend(reps)

    t_arr = np.array(unique_times, dtype=float)
    mean_arr = np.array([np.mean(reps_by_time[t]) for t in unique_times], dtype=float)
    std_arr = np.array([np.std(reps_by_time[t]) for t in unique_times], dtype=float)

    temp_time, temp_val = [], []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        t, temp = row[5], row[6]
        if t is None or temp is None:
            continue
        temp_time.append(float(t))
        temp_val.append(float(temp))

    return {
        'time': t_arr,
        'cfu_mean': mean_arr,
        'cfu_std': std_arr,
        'reps_by_time': reps_by_time,
        'temp_time': np.array(temp_time, dtype=float),
        'temp_C': np.array(temp_val, dtype=float),
    }


def load_sigma_knots_from_dir(outdir):
    temps = [4, 8, 15, 20, 25, 35, 37, 40]
    knots = {}
    for t in temps:
        path = os.path.join(outdir, f'samples_{t}C.npz')
        if not os.path.exists(path):
            raise FileNotFoundError(path)
        z = np.load(path)
        knots[float(t)] = float(z['sigma_MAP'])
    return knots


# ── Log-likelihood ───────────────────────────────────────────────────────────

def make_log_likelihood(exp_time, exp_cfu, exp_std, n_steps, time_to_step=10.0,
                        Kp=1e-4, eta=1.0, eta_phi=1.0, c=100.0, dynamic_c=True,
                        c_scale=10.0, dt=1e-4, n_newton=6):
    """
    Build log-likelihood closure for 1-param (sigma) estimation.

    Profiled likelihood: analytically optimize affine map a*phibar + b = cfu,
    then evaluate Gaussian log-likelihood with measurement noise.

    Parameters
    ----------
    dynamic_c : bool
        If True, use c(t) = c_scale*(1 - phi*psi) (dynamic nutrient coupling).
        If False, use fixed c (Felix static formulation with c=100).
    c_scale : float
        Scale for dynamic c. Default 10.0 (Felix: c(t) = 10*(1-phi*psi)).
    """
    exp_time_jax = jnp.array(exp_time)
    exp_cfu_jax = jnp.array(exp_cfu)
    exp_std_jax = jnp.array(exp_std)
    sigma_obs = jnp.maximum(exp_std_jax, 0.1)  # floor noise at 0.1 log10(CFU)

    def log_likelihood(theta):
        sigma = theta[0]

        # Run ODE
        if dynamic_c:
            phibar = simulate_monospecies_phibar_hist_dynamic_c(
                sigma, n_steps, dt=dt, Kp=Kp, eta=eta, eta_phi=eta_phi, c_scale=c_scale, n_newton=n_newton
            )
        else:
            phibar = simulate_monospecies_phibar_hist(
                sigma, n_steps, dt=dt, Kp=Kp, eta=eta, eta_phi=eta_phi, c=c, n_newton=n_newton
            )

        # Extract at observation times
        step_idx = jnp.clip((exp_time_jax * time_to_step).astype(int), 0, n_steps)
        sim_vals = phibar[step_idx]

        # Profile out affine mapping: cfu ≈ a*phibar + b (least squares)
        X = jnp.column_stack([sim_vals, jnp.ones_like(sim_vals)])
        coeff, _, _, _ = jnp.linalg.lstsq(X, exp_cfu_jax, rcond=None)
        predicted = X @ coeff

        # Gaussian log-likelihood
        residuals = (predicted - exp_cfu_jax) / sigma_obs
        logL = -0.5 * jnp.sum(residuals**2) - jnp.sum(jnp.log(sigma_obs))
        return logL

    return log_likelihood


def make_log_likelihood_dy_k(
    dy_time,
    dy_cfu,
    dy_std,
    dy_temp_time,
    dy_temp_C,
    sigma_knots,
    n_steps,
    time_to_step=10.0,
    Kp=1e-4,
    eta=1.0,
    eta_phi=1.0,
    dynamic_c=True,
    c_scale=10.0,
    dt=1e-4,
    n_newton=6,
):
    dy_time_jax = jnp.array(dy_time, dtype=jnp.float64)
    dy_cfu_jax = jnp.array(dy_cfu, dtype=jnp.float64)
    dy_std_jax = jnp.array(dy_std, dtype=jnp.float64)
    sigma_obs = jnp.maximum(dy_std_jax, 0.1)

    t_max = float(max(np.max(dy_time), np.max(dy_temp_time)))
    n_steps_from_time = int(np.ceil(t_max * time_to_step))
    if n_steps != n_steps_from_time:
        raise ValueError(f"n_steps mismatch: got {n_steps}, expected {n_steps_from_time} from time grid")

    t_grid = np.arange(n_steps + 1, dtype=float) / float(time_to_step)
    T_grid = np.interp(t_grid, np.asarray(dy_temp_time, dtype=float), np.asarray(dy_temp_C, dtype=float))

    temps = np.array(sorted(sigma_knots.keys()), dtype=float)
    sigmas = np.array([sigma_knots[t] for t in temps], dtype=float)
    sigma_base = np.interp(T_grid[:-1], temps, sigmas, left=sigmas[0], right=sigmas[-1])
    sigma_base_jax = jnp.array(sigma_base, dtype=jnp.float64)

    def log_likelihood(theta):
        k = theta[0]
        sigma_series = jnp.clip(k * sigma_base_jax, 1.0, 500.0)

        phibar = simulate_monospecies_phibar_hist_dynamic_c_sigma_series(
            sigma_series,
            n_steps,
            dt=dt,
            Kp=Kp,
            eta=eta,
            eta_phi=eta_phi,
            c_scale=c_scale,
            n_newton=n_newton,
        ) if dynamic_c else simulate_monospecies_phibar_hist(
            jnp.mean(sigma_series),
            n_steps,
            dt=dt,
            Kp=Kp,
            eta=eta,
            eta_phi=eta_phi,
            c=100.0,
            n_newton=n_newton,
        )

        step_idx = jnp.clip((dy_time_jax * time_to_step).astype(int), 0, n_steps)
        sim_vals = phibar[step_idx]

        X = jnp.column_stack([sim_vals, jnp.ones_like(sim_vals)])
        coeff, _, _, _ = jnp.linalg.lstsq(X, dy_cfu_jax, rcond=None)
        predicted = X @ coeff

        residuals = (predicted - dy_cfu_jax) / sigma_obs
        logL = -0.5 * jnp.sum(residuals**2) - jnp.sum(jnp.log(sigma_obs))
        return logL

    return log_likelihood


# ── Run one condition ────────────────────────────────────────────────────────

def run_tmcmc_condition(temp, data, n_particles=200, mutation='rw', seed=42,
                        max_stages=20, verbose=True, dynamic_c=False, c_scale=10.0,
                        dt=1e-4, n_newton=6, target_ess_ratio=0.5):
    """Run TMCMC for one temperature condition."""
    from tmcmc_nuts_engine import tmcmc_engine

    max_steps_map = {4: 4000, 8: 2000, 15: 1000, 20: 500,
                     25: 300, 35: 200, 37: 200, 40: 450}

    d = data[temp]
    n_steps = max_steps_map[temp]

    log_lik = make_log_likelihood(
        d['time'], d['cfu_mean'],
        d.get('cfu_std', np.full_like(d['cfu_mean'], 0.2)),
        n_steps, dynamic_c=dynamic_c, c_scale=c_scale, dt=dt, n_newton=n_newton,
    )

    # Per-temperature prior centered on Felix reference values.
    # Width: ±1 order of magnitude (factor 10) on each side, clipped to [1, 500].
    sigma_ref = FELIX_SIGMA[temp]
    lo = max(1.0, sigma_ref * 0.1)
    hi = min(500.0, sigma_ref * 10.0)
    prior_bounds = np.array([[lo, hi]])

    if verbose:
        mode_str = f"dynamic_c (scale={c_scale})" if dynamic_c else "static_c=100"
        print(f"\n{'='*60}")
        print(f"  {temp}°C: n_steps={n_steps}, n_obs={len(d['time'])}")
        print(f"  mutation={mutation}, n_particles={n_particles}, c_mode={mode_str}")
        print(f"  solver: dt={dt:g}, n_newton={n_newton}")
        print(f"  prior: sigma ∈ [{lo:.2f}, {hi:.2f}]  (felix_ref={sigma_ref})")
        print(f"{'='*60}")

    result = tmcmc_engine(
        log_likelihood=log_lik,
        prior_bounds=prior_bounds,
        mutation=mutation,
        n_particles=n_particles,
        max_stages=max_stages,
        target_ess_ratio=target_ess_ratio,
        hmc_step_size=0.05,
        hmc_n_leapfrog=5,
        nuts_max_depth=4,
        warmup_stages=2,
        seed=seed,
        label=f"{temp}C-{mutation}-{'dync' if dynamic_c else 'staticc'}",
        verbose=verbose,
    )

    sigma_MAP = float(result['theta_MAP'][0])
    samples = result['samples'][:, 0]
    sigma_mean = float(np.mean(samples))
    sigma_std = float(np.std(samples))
    sigma_ci = (float(np.percentile(samples, 2.5)),
                float(np.percentile(samples, 97.5)))

    if verbose:
        print(f"  MAP σ = {sigma_MAP:.2f}")
        print(f"  Mean σ = {sigma_mean:.2f} ± {sigma_std:.2f}")
        print(f"  95% CI: [{sigma_ci[0]:.2f}, {sigma_ci[1]:.2f}]")

    return {
        'temp': temp,
        'sigma_MAP': sigma_MAP,
        'sigma_mean': sigma_mean,
        'sigma_std': sigma_std,
        'sigma_ci_lo': sigma_ci[0],
        'sigma_ci_hi': sigma_ci[1],
        'n_stages': result['n_stages'],
        'total_time': result['total_time'],
        'samples': samples,
        'log_likelihoods': result['log_likelihoods'],
    }


def run_tmcmc_dy(
    dy,
    n_particles=400,
    mutation='nuts',
    seed=42,
    max_stages=30,
    verbose=True,
    dynamic_c=True,
    c_scale=10.0,
    dt=1e-4,
    n_newton=6,
    time_to_step=10.0,
    sigma_knots=None,
):
    from tmcmc_nuts_engine import tmcmc_engine

    if sigma_knots is None:
        sigma_knots = {float(t): float(FELIX_SIGMA[int(t)]) for t in [4, 8, 15, 20, 25, 35, 37, 40]}

    t_max = float(max(np.max(dy['time']), np.max(dy['temp_time'])))
    n_steps = int(np.ceil(t_max * time_to_step))

    log_lik = make_log_likelihood_dy_k(
        dy_time=dy['time'],
        dy_cfu=dy['cfu_mean'],
        dy_std=dy.get('cfu_std', np.full_like(dy['cfu_mean'], 0.2)),
        dy_temp_time=dy['temp_time'],
        dy_temp_C=dy['temp_C'],
        sigma_knots=sigma_knots,
        n_steps=n_steps,
        time_to_step=time_to_step,
        dynamic_c=dynamic_c,
        c_scale=c_scale,
        dt=dt,
        n_newton=n_newton,
    )

    prior_bounds = np.array([[0.05, 5.0]], dtype=float)

    if verbose:
        mode_str = f"dynamic_c (scale={c_scale})" if dynamic_c else "static_c=100"
        print(f"\n{'='*60}")
        print("  Dy (dynamic temperature) TMCMC")
        print(f"  n_steps={n_steps}, n_obs={len(dy['time'])}")
        print(f"  mutation={mutation}, n_particles={n_particles}, c_mode={mode_str}")
        print(f"  solver: dt={dt:g}, n_newton={n_newton}, time_to_step={time_to_step:g}")
        print("  parameter: k (sigma(t)=k*interp(sigma_knots,T(t)))")
        print(f"  prior: k ∈ [{prior_bounds[0,0]:.3g}, {prior_bounds[0,1]:.3g}]")
        print(f"{'='*60}")

    result = tmcmc_engine(
        log_likelihood=log_lik,
        prior_bounds=prior_bounds,
        mutation=mutation,
        n_particles=n_particles,
        max_stages=max_stages,
        target_ess_ratio=0.5,
        hmc_step_size=0.05,
        hmc_n_leapfrog=5,
        nuts_max_depth=6,
        warmup_stages=2,
        seed=seed,
        label=f"Dy-{mutation}-{'dync' if dynamic_c else 'staticc'}",
        verbose=verbose,
    )

    k_MAP = float(result['theta_MAP'][0])
    samples = result['samples'][:, 0]
    k_mean = float(np.mean(samples))
    k_std = float(np.std(samples))
    k_ci = (float(np.percentile(samples, 2.5)), float(np.percentile(samples, 97.5)))

    if verbose:
        print(f"  MAP k = {k_MAP:.4f}")
        print(f"  Mean k = {k_mean:.4f} ± {k_std:.4f}")
        print(f"  95% CI: [{k_ci[0]:.4f}, {k_ci[1]:.4f}]")

    return {
        'k_MAP': k_MAP,
        'k_mean': k_mean,
        'k_std': k_std,
        'k_ci_lo': k_ci[0],
        'k_ci_hi': k_ci[1],
        'n_stages': result['n_stages'],
        'total_time': result['total_time'],
        'samples': samples,
        'log_likelihoods': result['log_likelihoods'],
        'n_steps': n_steps,
        'time_to_step': float(time_to_step),
    }


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--temp', type=int, default=None,
                        help='Single temperature (4,8,15,20,25,35,37,40). Default: all.')
    parser.add_argument('--device', type=str, default='auto', choices=['auto', 'gpu', 'cpu'],
                        help='JAX device selection. Must be set before JAX import; this flag is read early.')
    parser.add_argument('--mutation', type=str, default='rw', choices=['rw', 'nuts', 'hmc'])
    parser.add_argument('--n_particles', type=int, default=200)
    parser.add_argument('--max_stages', type=int, default=30)
    parser.add_argument('--seed', type=int, default=42)
    parser.add_argument('--outdir', type=str, default=None,
                        help='Output directory. Default: _tmcmc_results_staticc or _tmcmc_results_dyncc')
    parser.add_argument('--use-dy', action='store_true',
                        help='Run TMCMC on Dy (dynamic temperature) data instead of St.')
    parser.add_argument('--dy-sheet', type=str, default='Dy')
    parser.add_argument('--dy-time-to-step', type=float, default=10.0,
                        help='Mapping from hours to step index for Dy. Default 10.0')
    parser.add_argument('--sigma-knots-dir', type=str, default=None,
                        help='Directory containing samples_{T}C.npz to build sigma_knots for Dy.')
    parser.add_argument('--dynamic-c', action='store_true',
                        help='Use dynamic c(t)=10*(1-phi*psi) instead of fixed c=100')
    parser.add_argument('--c-scale', type=float, default=10.0,
                        help='Scale for dynamic c. Default 10.0')
    parser.add_argument('--dt', type=float, default=1e-4,
                        help='Implicit Euler time step for solver. Default 1e-4')
    parser.add_argument('--n-newton', type=int, default=6,
                        help='Newton iterations per time step. Default 6')
    parser.add_argument('--target-ess-ratio', type=float, default=0.5,
                        help='Target ESS ratio for TMCMC tempering. Default 0.5')
    args = parser.parse_args()
    args.dynamic_c = args.dynamic_c  # rename for clarity

    try:
        backend = jax.default_backend()
        devs = ", ".join([f"{d.platform}:{d.device_kind}" for d in jax.devices()])
        print(f"JAX backend: {backend} | devices: {devs}")
    except Exception:
        pass

    if args.outdir is None:
        if args.use_dy:
            args.outdir = '_tmcmc_results_dy'
        else:
            args.outdir = '_tmcmc_results_dyncc' if args.dynamic_c else '_tmcmc_results_staticc'

    xlsx_path = os.path.join(os.path.dirname(__file__), 'raw data.xlsx')

    if args.use_dy:
        os.makedirs(args.outdir, exist_ok=True)

        dy = load_dynamic_data_with_reps(xlsx_path, sheet=args.dy_sheet)
        if args.sigma_knots_dir is not None:
            sigma_knots = load_sigma_knots_from_dir(args.sigma_knots_dir)
        else:
            sigma_knots = {float(t): float(FELIX_SIGMA[int(t)]) for t in [4, 8, 15, 20, 25, 35, 37, 40]}

        t_total = time.time()
        res = run_tmcmc_dy(
            dy,
            n_particles=args.n_particles,
            mutation=args.mutation,
            seed=args.seed,
            max_stages=args.max_stages,
            dynamic_c=True,
            c_scale=args.c_scale,
            dt=args.dt,
            n_newton=args.n_newton,
            time_to_step=args.dy_time_to_step,
            sigma_knots=sigma_knots,
        )
        elapsed_total = time.time() - t_total

        np.savez(
            os.path.join(args.outdir, 'samples_Dy_k.npz'),
            samples=res['samples'],
            log_likelihoods=res['log_likelihoods'],
            k_MAP=res['k_MAP'],
            n_steps=res['n_steps'],
            time_to_step=res['time_to_step'],
        )

        csv_path = os.path.join(args.outdir, 'tmcmc_summary_Dy_k.csv')
        with open(csv_path, 'w', newline='') as f:
            w = csv.writer(f)
            w.writerow(['k_MAP', 'k_mean', 'k_std', 'CI_lo_2.5', 'CI_hi_97.5', 'n_stages', 'time_s', 'n_steps', 'time_to_step'])
            w.writerow([f"{res['k_MAP']:.6f}", f"{res['k_mean']:.6f}", f"{res['k_std']:.6f}",
                        f"{res['k_ci_lo']:.6f}", f"{res['k_ci_hi']:.6f}", res['n_stages'],
                        f"{res['total_time']:.1f}", res['n_steps'], f"{res['time_to_step']:.6g}"])
        print(f"Saved: {csv_path}")

        json_path = os.path.join(args.outdir, 'tmcmc_summary_Dy_k.json')
        summary = {
            'mode': 'Dy_k',
            'mutation': args.mutation,
            'n_particles': args.n_particles,
            'dynamic_c': True,
            'c_scale': args.c_scale,
            'dt': args.dt,
            'n_newton': args.n_newton,
            'dy_sheet': args.dy_sheet,
            'dy_time_to_step': args.dy_time_to_step,
            'sigma_knots_dir': args.sigma_knots_dir,
            'k_MAP': res['k_MAP'],
            'k_mean': res['k_mean'],
            'k_std': res['k_std'],
            'k_ci_lo': res['k_ci_lo'],
            'k_ci_hi': res['k_ci_hi'],
            'n_stages': res['n_stages'],
            'total_time': elapsed_total,
        }
        with open(json_path, 'w') as f:
            json.dump(summary, f, indent=2)
        print(f"Saved: {json_path}")
        return

    data = load_static_data(xlsx_path)

    temps = [args.temp] if args.temp else [4, 8, 15, 20, 25, 35, 37, 40]

    os.makedirs(args.outdir, exist_ok=True)

    # paper_sigma_miss = {4: 25, 8: 50, 15: 50, 20: 70, 25: 70, 35: 110, 37: 110, 40: 50}
    paper_sigma = {4: 4.25, 8: 10.0, 15: 25.0, 20: 50.0, 25: 100.0, 35: 110.0, 37: 115.0, 40: 40.0}
    
    all_results = []
    t_total = time.time()

    for temp in temps:
        res = run_tmcmc_condition(
            temp, data,
            n_particles=args.n_particles,
            mutation=args.mutation,
            seed=args.seed,
            max_stages=args.max_stages,
            dynamic_c=args.dynamic_c,
            c_scale=args.c_scale,
            dt=args.dt,
            n_newton=args.n_newton,
            target_ess_ratio=args.target_ess_ratio,
        )
        all_results.append(res)

        # Save samples per condition
        np.savez(
            os.path.join(args.outdir, f'samples_{temp}C.npz'),
            samples=res['samples'],
            log_likelihoods=res['log_likelihoods'],
            sigma_MAP=res['sigma_MAP'],
        )

    elapsed_total = time.time() - t_total

    # ── Summary table ────────────────────────────────────────────────────
    print(f"\n{'='*90}")
    c_mode = f"dynamic_c(scale={args.c_scale})" if args.dynamic_c else "static_c=100"
    print(f"TMCMC Monospecies Summary ({args.mutation.upper()}, {args.n_particles}p, {c_mode})")
    print(f"{'='*90}")
    print(f"{'Temp':>6s} | {'Paper σ':>8s} | {'MAP σ':>8s} | {'Mean±Std':>14s} | "
          f"{'95% CI':>16s} | {'Stages':>6s} | {'Time':>6s}")
    print('-' * 90)

    for res in all_results:
        t = res['temp']
        print(f"{t:>4d}°C | {paper_sigma[t]:>8.2f} | {res['sigma_MAP']:>8.2f} | "
              f"{res['sigma_mean']:>6.2f} ± {res['sigma_std']:>5.2f} | "
              f"[{res['sigma_ci_lo']:>6.2f}, {res['sigma_ci_hi']:>6.2f}] | "
              f"{res['n_stages']:>6d} | {res['total_time']:>5.1f}s")

    print(f"\nTotal elapsed: {elapsed_total:.1f}s")

    # ── Save summary CSV ─────────────────────────────────────────────────
    csv_path = os.path.join(args.outdir, 'tmcmc_summary.csv')
    with open(csv_path, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['Temperature_C', 'paper_sigma', 'MAP_sigma', 'mean_sigma',
                     'std_sigma', 'CI_lo_2.5', 'CI_hi_97.5', 'n_stages', 'time_s'])
        for res in all_results:
            w.writerow([res['temp'], paper_sigma[res['temp']],
                         f"{res['sigma_MAP']:.4f}", f"{res['sigma_mean']:.4f}",
                         f"{res['sigma_std']:.4f}", f"{res['sigma_ci_lo']:.4f}",
                         f"{res['sigma_ci_hi']:.4f}", res['n_stages'],
                         f"{res['total_time']:.1f}"])
    print(f"Saved: {csv_path}")

    # ── Save summary JSON ────────────────────────────────────────────────
    json_path = os.path.join(args.outdir, 'tmcmc_summary.json')
    summary = {
        'mutation': args.mutation,
        'n_particles': args.n_particles,
        'dynamic_c': args.dynamic_c,
        'c_scale': args.c_scale,
        'dt': args.dt,
        'n_newton': args.n_newton,
        'total_time': elapsed_total,
        'conditions': [{k: v for k, v in res.items()
                        if k not in ('samples', 'log_likelihoods')}
                       for res in all_results],
    }
    with open(json_path, 'w') as f:
        json.dump(summary, f, indent=2)
    print(f"Saved: {json_path}")

    # ── XLSX ─────────────────────────────────────────────────────────────
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Summary'
        ws.append(['Temperature (°C)', 'Paper σ', 'MAP σ', 'Mean σ', 'Std σ',
                    '95% CI lo', '95% CI hi', 'Stages', 'Time (s)'])
        for res in all_results:
            ws.append([res['temp'], paper_sigma[res['temp']],
                       round(res['sigma_MAP'], 4), round(res['sigma_mean'], 4),
                       round(res['sigma_std'], 4), round(res['sigma_ci_lo'], 4),
                       round(res['sigma_ci_hi'], 4), res['n_stages'],
                       round(res['total_time'], 1)])

        # Posterior samples sheets
        for res in all_results:
            ws_s = wb.create_sheet(f"{res['temp']}C_posterior")
            ws_s.append(['sample_idx', 'sigma', 'log_likelihood'])
            for i, (s, ll) in enumerate(zip(res['samples'], res['log_likelihoods'])):
                ws_s.append([i, round(float(s), 6), round(float(ll), 4)])

        xlsx_out = os.path.join(args.outdir, 'tmcmc_summary.xlsx')
        wb.save(xlsx_out)
        print(f"Saved: {xlsx_out}")
    except ImportError:
        print("openpyxl not available, xlsx skipped")

if __name__ == '__main__':
    main()
