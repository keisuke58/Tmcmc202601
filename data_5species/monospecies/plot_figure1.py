import os
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker

# Configure matplotlib for high quality output
plt.rcParams.update({
    'font.family': 'sans-serif',
    'font.sans-serif': ['Arial', 'Helvetica', 'DejaVu Sans'],
    'font.size': 12,
    'axes.labelsize': 14,
    'axes.titlesize': 14,
    'xtick.labelsize': 12,
    'ytick.labelsize': 12,
    'legend.fontsize': 12,
    'figure.dpi': 300,
    'axes.linewidth': 1.5,
    'xtick.major.width': 1.5,
    'ytick.major.width': 1.5,
})

def load_dynamic_data(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Dy']
    
    times, reps1, reps2, reps3 = [], [], [], []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None: continue
        times.append(float(row[0]))
        reps1.append(float(row[1]) if row[1] is not None else np.nan)
        reps2.append(float(row[2]) if row[2] is not None else np.nan)
        reps3.append(float(row[3]) if row[3] is not None else np.nan)
        
    times = np.array(times)
    reps = np.column_stack([reps1, reps2, reps3])
    
    # Calculate means and stds ignoring NaNs
    means = np.nanmean(reps, axis=1)
    stds = np.nanstd(reps, axis=1)
    
    # Load temperature profile
    temp_time, temp_C = [], []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if len(row) > 5 and row[5] is not None and row[6] is not None:
            temp_time.append(float(row[5]))
            temp_C.append(float(row[6]))
            
    return times, reps, means, stds, np.array(temp_time), np.array(temp_C)

def plot_figure1():
    xlsx_path = os.path.join(os.path.dirname(__file__), 'raw data.xlsx')
    times, reps, means, stds, temp_time, temp_C = load_dynamic_data(xlsx_path)
    
    fig, (ax_temp, ax_cfu) = plt.subplots(2, 1, figsize=(10, 8), height_ratios=[1, 2], sharex=True)
    fig.subplots_adjust(hspace=0.05)
    
    # 1. Plot Temperature Profile
    ax_temp.plot(temp_time, temp_C, '-', color='#E63946', linewidth=2.5, label='Temperature')
    ax_temp.set_ylabel('Temperature (°C)', color='#E63946', fontweight='bold')
    ax_temp.tick_params(axis='y', colors='#E63946')
    ax_temp.set_ylim(-40, 60)
    ax_temp.grid(True, linestyle='--', alpha=0.5)
    
    # Add temperature markers/annotations for clarity
    peaks = [(3.0, 37.0), (10.0, 40.0)]
    for tx, ty in peaks:
        ax_temp.plot(tx, ty, 'o', color='#E63946', markersize=6)
    
    # 2. Plot CFU Data
    # Plot individual replicates
    colors = ['#457B9D', '#1D3557', '#A8DADC']
    markers = ['o', 's', '^']
    
    for i in range(3):
        valid = ~np.isnan(reps[:, i])
        ax_cfu.plot(times[valid], reps[valid, i], marker=markers[i], linestyle='', 
                    color=colors[0], alpha=0.5, markersize=7, 
                    label=f'Replicate {i+1}' if i==0 else "")
    
    # Plot mean and error bars
    ax_cfu.errorbar(times, means, yerr=stds, fmt='D', color='#1D3557', 
                    linewidth=2, capsize=5, capthick=1.5, markersize=8,
                    label='Mean $\pm$ SD', zorder=5)
    
    # Formatting
    ax_cfu.set_xlabel('Time (days)', fontweight='bold')
    ax_cfu.set_ylabel('$\log_{10}$(CFU/g)', fontweight='bold')
    ax_cfu.set_ylim(3, 9)
    ax_cfu.grid(True, linestyle='--', alpha=0.5)
    
    # Add minor ticks
    ax_cfu.xaxis.set_minor_locator(ticker.AutoMinorLocator())
    ax_cfu.yaxis.set_minor_locator(ticker.AutoMinorLocator())
    ax_temp.yaxis.set_minor_locator(ticker.AutoMinorLocator())
    
    # Legends
    ax_cfu.legend(loc='lower right', frameon=True, fancybox=True, shadow=True)
    
    # Title
    fig.suptitle('Dynamic Temperature Profile and Biofilm Growth', 
                 fontsize=16, fontweight='bold', y=0.95)
    
    # Save high quality plots
    out_dir = os.path.join(os.path.dirname(__file__), 'figures')
    os.makedirs(out_dir, exist_ok=True)
    
    plt.savefig(os.path.join(out_dir, 'figure1_raw_data.png'), bbox_inches='tight', dpi=300)
    plt.savefig(os.path.join(out_dir, 'figure1_raw_data.pdf'), bbox_inches='tight')
    plt.close()
    
    print("Saved to:", os.path.join(out_dir, 'figure1_raw_data.png'))

def plot_figure1_overlay():
    xlsx_path = os.path.join(os.path.dirname(__file__), 'raw data.xlsx')
    times, reps, means, stds, temp_time, temp_C = load_dynamic_data(xlsx_path)
    
    fig, ax_cfu = plt.subplots(figsize=(10, 6))
    ax_temp = ax_cfu.twinx()
    
    # 1. Plot Temperature Profile (Right Axis)
    ax_temp.plot(temp_time, temp_C, '-', color='#E63946', linewidth=2.5, alpha=0.8, label='Temperature')
    ax_temp.set_ylabel('Temperature (°C)', color='#E63946', fontweight='bold')
    ax_temp.tick_params(axis='y', colors='#E63946')
    ax_temp.set_ylim(-40, 60)
    
    # 2. Plot CFU Data (Left Axis)
    colors = ['#457B9D', '#1D3557', '#A8DADC']
    markers = ['o', 's', '^']
    
    for i in range(3):
        valid = ~np.isnan(reps[:, i])
        ax_cfu.plot(times[valid], reps[valid, i], marker=markers[i], linestyle='', 
                    color=colors[0], alpha=0.5, markersize=7, 
                    label=f'Replicate {i+1}' if i==0 else "")
    
    ax_cfu.errorbar(times, means, yerr=stds, fmt='D', color='#1D3557', 
                    linewidth=2, capsize=5, capthick=1.5, markersize=8,
                    label='Mean $\pm$ SD', zorder=5)
    
    # Formatting
    ax_cfu.set_xlabel('Time (days)', fontweight='bold')
    ax_cfu.set_ylabel('$\log_{10}$(CFU/g)', color='#1D3557', fontweight='bold')
    ax_cfu.tick_params(axis='y', colors='#1D3557')
    ax_cfu.set_ylim(3, 9)
    ax_cfu.grid(True, linestyle='--', alpha=0.5)
    
    # Add minor ticks
    ax_cfu.xaxis.set_minor_locator(ticker.AutoMinorLocator())
    ax_cfu.yaxis.set_minor_locator(ticker.AutoMinorLocator())
    ax_temp.yaxis.set_minor_locator(ticker.AutoMinorLocator())
    
    # Combined Legend
    lines_cfu, labels_cfu = ax_cfu.get_legend_handles_labels()
    lines_temp, labels_temp = ax_temp.get_legend_handles_labels()
    ax_cfu.legend(lines_cfu + lines_temp, labels_cfu + labels_temp, 
                  loc='lower right', frameon=True, fancybox=True, shadow=True)
    
    # Title
    plt.title('Dynamic Temperature Profile and Biofilm Growth (Overlay)', 
              fontsize=16, fontweight='bold', pad=15)
    
    # Save high quality plots
    out_dir = os.path.join(os.path.dirname(__file__), 'figures')
    os.makedirs(out_dir, exist_ok=True)
    
    plt.savefig(os.path.join(out_dir, 'figure1_raw_data_overlay.png'), bbox_inches='tight', dpi=300)
    plt.savefig(os.path.join(out_dir, 'figure1_raw_data_overlay.pdf'), bbox_inches='tight')
    plt.close()
    
    print("Saved to:", os.path.join(out_dir, 'figure1_raw_data_overlay.png'))

if __name__ == '__main__':
    plot_figure1()
    plot_figure1_overlay()
